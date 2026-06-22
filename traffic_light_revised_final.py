import os
import sys
import subprocess
import tempfile
from typing import List, Tuple, Dict
import numpy as np
import openpyxl
import matplotlib.colors as mcolors
import plotly.graph_objects as go
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
)
from reportlab.pdfgen import canvas

# -------------------------------------------------------------------
# Font Registration
# -------------------------------------------------------------------
try:
    pdfmetrics.registerFont(TTFont('Optima', '/System/Library/Fonts/Optima.ttc'))
    pdfmetrics.registerFont(TTFont('Optima-Bold', '/System/Library/Fonts/Optima.ttc'))
except Exception:
    pass

# -------------------------------------------------------------------
# macOS Native File Dialogs
# -------------------------------------------------------------------
def macos_choose_file(prompt="Select an Excel workbook") -> str:
    script = f'''try
  set theFile to (choose file with prompt "{prompt}")
  POSIX path of theFile
on error
  return ""
end try'''
    proc = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    path = proc.stdout.strip()
    return path if path else None

def macos_save_dialog(default_name="STAJ_Report.pdf", prompt="Save PDF Report As") -> str:
    safe_name = os.path.basename(default_name)
    script = f'''try
  set theFile to (choose file name with prompt "{prompt}" default name "{safe_name}")
  POSIX path of theFile
on error
  return ""
end try'''
    proc = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    path = proc.stdout.strip()
    return path if path else None

# -------------------------------------------------------------------
# Utility Helpers
# -------------------------------------------------------------------
def is_number(value) -> bool:
    if value in (None, "", "-", " "):
        return False
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False

from typing import List

def extract_units(unit_text: str, all_unique_units: List[str]) -> List[str]:
    if not unit_text:
        return []

    s = str(unit_text).strip()

    # Expand ALL UNITS to all discovered units
    if s.lower() == "all units":
        return all_unique_units

    # Normalize separators
    for sep in [';', '/', '&', ' and ']:
        s = s.replace(sep, ',')

    return [p.strip() for p in s.split(',') if p.strip()]

def get_smooth_steps(num_steps=50) -> List[Dict]:
    cmap = mcolors.LinearSegmentedColormap.from_list("rag_smooth", ["#d62728", "#FFD700", "#006400"])
    steps = []
    for i in range(num_steps):
        color = mcolors.to_hex(cmap(i / (num_steps - 1)))
        start = (i / num_steps) * 100
        end = ((i + 1) / num_steps) * 100
        steps.append({'range': [start, end], 'color': color})
    return steps

def get_bar_color(val: float) -> str:
    v = max(0.0, min(val, 100.0))
    cmap = mcolors.LinearSegmentedColormap.from_list("rag_smooth", ["#d62728", "#FFD700", "#006400"])
    return mcolors.to_hex(cmap(v / 100.0))

# -------------------------------------------------------------------
# Exact-Scaled Segmented Progress Bar with In-Bar Labels
# -------------------------------------------------------------------
def create_segmented_progress_bar(q_pcts: List[float], total_width: float = 400.0) -> Table:
    increments = []
    prev = 0.0
    for val in q_pcts:
        val = max(0.0, min(val, 100.0))
        inc = max(0.0, val - prev)
        increments.append(inc)
        prev = val

    overall_pct = sum(increments)
    remaining_pct = max(0.0, 100.0 - overall_pct)
    
    col_widths = []
    background_colors = []
    cell_contents = []
    
    styles = getSampleStyleSheet()
    in_bar_style = ParagraphStyle(
        name="InBarStyle", 
        parent=styles["Normal"], 
        fontName="Optima-Bold", 
        fontSize=6.5, 
        leading=8, 
        alignment=1, 
        textColor=colors.black
    )

    for i, inc in enumerate(increments):
        if inc > 0:
            width = (inc / 100.0) * total_width
            col_widths.append(width)
            background_colors.append(get_bar_color(q_pcts[i]))
            
            if width > 25:
                p_text = f"Q{i+1}: {q_pcts[i]:.1f}%"
                cell_contents.append(Paragraph(p_text, in_bar_style))
            else:
                cell_contents.append("")
            
    if remaining_pct > 0:
        col_widths.append((remaining_pct / 100.0) * total_width)
        background_colors.append("#EAEAEA")
        cell_contents.append("")

    if not col_widths:
        col_widths = [total_width]
        background_colors = ["#EAEAEA"]
        cell_contents = [""]

    bar_data = [cell_contents]
    bar_table = Table(bar_data, colWidths=col_widths, rowHeights=15)
    
    t_styles = [
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
    ]
    
    for idx, hex_color in enumerate(background_colors):
        t_styles.append(('BACKGROUND', (idx, 0), (idx, 0), colors.HexColor(hex_color)))
        
    bar_table.setStyle(TableStyle(t_styles))
    return bar_table

# -------------------------------------------------------------------
# Gauge Rendering
# -------------------------------------------------------------------
def draw_color_coded_gauge(value: float, out_path="gauge_dial.png"):
    display_value = max(0.0, min(value, 100.0))
    domain_x, domain_y = [0.0, 1.0], [0.0, 1.0]
    cx, cy = 0.5, 0.25 
    angle = (1 - display_value / 100.0) * np.pi
    radius = 0.40
    x_tip = cx + radius * np.cos(angle)
    y_tip = cy + radius * np.sin(angle)
    
    fig = go.Figure()
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=display_value,
        number={'suffix': "%", 'font': {'size': 60}, 'valueformat': '.1f'},
        domain={'x': domain_x, 'y': domain_y},
        gauge={
            'axis': {
                'range': [0, 100],
                'tickmode': 'array',
                'tickvals': list(range(0, 101, 10)),
                'tickfont': {'size': 24},
            },
            'bar': {'color': "black", 'thickness': 0.1},
            'steps': get_smooth_steps(num_steps=50)
        }
    ))
    fig.add_shape(type="line", x0=cx, y0=cy, x1=x_tip, y1=y_tip,
                  xref='paper', yref='paper', line=dict(color="black", width=4), layer='above')
    fig.add_shape(type="circle", x0=cx - 0.02, y0=cy - 0.02, x1=cx + 0.02, y1=cy + 0.02,
                  xref='paper', yref='paper', fillcolor="black", line_color="black", layer='above')
    fig.update_layout(width=650, height=600, paper_bgcolor="white", margin=dict(l=30, r=30, t=70, b=30))
    fig.write_image(out_path, width=650, height=420)
    return fig

# -------------------------------------------------------------------
# Reporting Period Detection (Identifies active FY and quarter)
# -------------------------------------------------------------------
def get_active_year_start_col(ws, target_col=9) -> int:
    """Finds which FY contains the most recent data entry."""
    col_map = {
        10: "FY 2024/2025", 14: "FY 2025/2026", 18: "FY 2026/2027"
    }
    last_col_idx = 9
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=target_col).value in (None, ""):
            continue
        for c in range(9, 21):
            val = ws.cell(row=row, column=c).value
            if val not in (None, "", "-", " "):
                if c > last_col_idx:
                    last_col_idx = c

    # Snap to the beginning column of that active FY range
    if 10 <= last_col_idx <= 12: return 10
    if 14 <= last_col_idx <= 16: return 14
    return 18

def detect_reporting_period(ws, target_col=9) -> str:
    col_map = {
        10:  ("FY 2024/2025", 1), 11: ("FY 2024/2025", 2), 12: ("FY 2024/2025", 3), 13: ("FY 2024/2025", 4),
        14: ("FY 2025/2026", 1), 15: ("FY 2025/2026", 2), 16: ("FY 2025/2026", 3), 17: ("FY 2025/2026", 4),
        18: ("FY 2026/2027", 1), 19: ("FY 2026/2027", 2), 20: ("FY 2026/2027", 3), 21: ("FY 2026/2027", 4),
    }

    last_col_index = -1
    last_filled = None

    for row in range(4, ws.max_row + 1):
        target_value = ws.cell(row=row, column=target_col).value
        if target_value in (None, ""):
            continue
        for col, (fy, qnum) in col_map.items():
            val = ws.cell(row=row, column=col).value
            if val not in (None, "", "-", " "):
                if col > last_col_index:
                    last_col_index = col
                    last_filled = (fy, qnum)

    if last_filled:
        fy, qnum = last_filled
        return f"Quarter {qnum} of {fy}"
    return "Reporting Period Undetermined"

# -------------------------------------------------------------------
# Precise Quarterly Isolator (Limits accumulation to active year only)
# -------------------------------------------------------------------
def get_quarterly_cumulative_percentages(ws, row_idx, target_col=9) -> List[float]:
    target_val = ws.cell(row=row_idx, column=target_col).value
    if not is_number(target_val) or float(target_val) <= 0:
        return [0.0, 0.0, 0.0, 0.0]

    target = float(target_val)
    # Detect which FY we should be analyzing
    start_col = get_active_year_start_col(ws, target_col)

    # Only step through the 4 quarters of that specific Fiscal Year
    cumulative_achieved = 0.0
    quarter_pcts = []

    for q_idx in range(4):
        c = start_col + q_idx
        v = ws.cell(row=row_idx, column=c).value
        if is_number(v):
            cumulative_achieved += float(v)
        
        pct = min((cumulative_achieved / target) * 100.0, 100.0)
        quarter_pcts.append(pct)

    return quarter_pcts

# -------------------------------------------------------------------
# Advanced Grouped Calculation Helpers
# -------------------------------------------------------------------
def compute_outcome_quarterly_progress(ws) -> Dict[str, Dict]:
    target_col = 9
    last_row = ws.max_row
    current_outcome = 1
    outcomes = {}
    indicator_q_pcts = []

    for i in range(4, last_row + 1):
        target_value = ws.cell(row=i, column=target_col).value

        if target_value in (None, ""):
            if len(indicator_q_pcts) > 0:
                q_avgs = [sum(q) / len(indicator_q_pcts) for q in zip(*indicator_q_pcts)]
                outcomes[f"Outcome {current_outcome}"] = {
                    "overall": q_avgs[3],
                    "quarters": q_avgs
                }
                current_outcome += 1
            indicator_q_pcts = []
            continue

        if not is_number(target_value):
            continue

        row_q_pcts = get_quarterly_cumulative_percentages(ws, i)
        indicator_q_pcts.append(row_q_pcts)

    if len(indicator_q_pcts) > 0:
        q_avgs = [sum(q) / len(indicator_q_pcts) for q in zip(*indicator_q_pcts)]
        outcomes[f"Outcome {current_outcome}"] = {
            "overall": q_avgs[3],
            "quarters": q_avgs
        }

    return outcomes



def compute_all_progress_old(ws):
    target_col = 9
    last_row = ws.max_row

    # Find last valid data row
    for r in range(ws.max_row, 0, -1):
        if is_number(ws.cell(row=r, column=target_col).value):
            last_row = r
            break

    # ------------------------------------------------------------------
    # FIRST PASS:
    # Build all unique units excluding "ALL UNITS"
    # ------------------------------------------------------------------
    all_units = set()

    for i in range(4, last_row + 1):
        unit_cell = ws.cell(row=i, column=7).value

        if unit_cell in (None, ""):
            continue

        s = str(unit_cell).strip()

        # Skip ALL UNITS for discovery
        if s.lower() == "all units":
            continue

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')

        units = [p.strip() for p in s.split(',') if p.strip()]
        all_units.update(units)

    all_unique_units = sorted(all_units)

    # ------------------------------------------------------------------
    # Helper function
    # ------------------------------------------------------------------
    def extract_units(unit_text):
        if not unit_text:
            return []

        s = str(unit_text).strip()

        # Expand ALL UNITS
        if s.lower() == "all units":
            return all_unique_units

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')

        return [p.strip() for p in s.split(',') if p.strip()]

    # ------------------------------------------------------------------
    # SECOND PASS:
    # Compute progress
    # ------------------------------------------------------------------
    row_percentages = []
    unit_q_sums = {}
    unit_q_cnts = {}

    for i in range(4, last_row + 1):

        target_val = ws.cell(row=i, column=target_col).value

        if not is_number(target_val) or float(target_val) == 0:
            continue

        row_pcts = get_quarterly_cumulative_percentages(ws, i)
        row_percentages.append(row_pcts[3])

        unit_cell = ws.cell(row=i, column=7).value

        if unit_cell not in (None, ""):

            units = extract_units(unit_cell)

            for u in units:

                if u not in unit_q_sums:
                    unit_q_sums[u] = [0.0, 0.0, 0.0, 0.0]
                    unit_q_cnts[u] = 0

                for q_idx in range(4):
                    unit_q_sums[u][q_idx] += row_pcts[q_idx]

                unit_q_cnts[u] += 1

    # ------------------------------------------------------------------
    # Overall progress
    # ------------------------------------------------------------------
    overall_pct = (
        sum(row_percentages) / len(row_percentages)
        if row_percentages else 0.0
    )

    # ------------------------------------------------------------------
    # Unit averages
    # ------------------------------------------------------------------
    unit_avgs = {}

    for u in unit_q_sums:

        cnt = unit_q_cnts[u]

        unit_avgs[u] = {
            "overall": unit_q_sums[u][3] / cnt if cnt > 0 else 0.0,
            "quarters": (
                [val / cnt for val in unit_q_sums[u]]
                if cnt > 0 else [0, 0, 0, 0]
            )
        }

    outcomes = compute_outcome_quarterly_progress(ws)

    unit_avgs = dict(
        sorted(
            unit_avgs.items(),
            key=lambda item: item[1]["overall"],
            reverse=True
        )
    )

    return overall_pct, outcomes, unit_avgs


def compute_all_progress_old2(ws):
    target_col = 9
    last_row = ws.max_row

    # Find last valid data row
    for r in range(ws.max_row, 0, -1):
        if is_number(ws.cell(row=r, column=target_col).value):
            last_row = r
            break

    # ------------------------------------------------------------------
    # FIRST PASS: Build all unique units excluding "ALL UNITS"
    # ------------------------------------------------------------------
    all_units = set()
    for i in range(4, last_row + 1):
        unit_cell = ws.cell(row=i, column=7).value
        if unit_cell in (None, ""):
            continue

        s = str(unit_cell).strip()
        if s.lower() == "all units":
            continue

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')

        units = [p.strip() for p in s.split(',') if p.strip()]
        all_units.update(units)

    all_unique_units = sorted(all_units)

    # Helper function
    def extract_units(unit_text):
        if not unit_text:
            return []
        s = str(unit_text).strip()
        if s.lower() == "all units":
            return all_unique_units

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')
        return [p.strip() for p in s.split(',') if p.strip()]

    # ------------------------------------------------------------------
    # SECOND PASS: Compute progress
    # ------------------------------------------------------------------
    row_percentages = []
    unit_q_sums = {}
    unit_q_cnts = {}

    for i in range(4, last_row + 1):
        target_val = ws.cell(row=i, column=target_col).value

        if not is_number(target_val) or float(target_val) == 0:
            continue

        row_pcts = get_quarterly_cumulative_percentages(ws, i)
        row_percentages.append(row_pcts[3])

        unit_cell = ws.cell(row=i, column=7).value
        if unit_cell not in (None, ""):
            units = extract_units(unit_cell)
            for u in units:
                if u not in unit_q_sums:
                    unit_q_sums[u] = [0.0, 0.0, 0.0, 0.0]
                    unit_q_cnts[u] = 0

                for q_idx in range(4):
                    unit_q_sums[u][q_idx] += row_pcts[q_idx]

                unit_q_cnts[u] += 1

    # ------------------------------------------------------------------
    # Overall progress (Rounded to 2 decimal places)
    # ------------------------------------------------------------------
    overall_pct = (
        round(sum(row_percentages) / len(row_percentages), 2)
        if row_percentages else 0.0
    )

    # ------------------------------------------------------------------
    # Unit averages (Rounded to 2 decimal places)
    # ------------------------------------------------------------------
    unit_avgs = {}
    for u in unit_q_sums:
        cnt = unit_q_cnts[u]
        unit_avgs[u] = {
            "overall": round(unit_q_sums[u][3] / cnt, 2) if cnt > 0 else 0.0,
            "quarters": (
                [round(val / cnt, 2) for val in unit_q_sums[u]]
                if cnt > 0 else [0.0, 0.0, 0.0, 0.0]
            )
        }

    outcomes = compute_outcome_quarterly_progress(ws)

    unit_avgs = dict(
        sorted(
            unit_avgs.items(),
            key=lambda item: item[1]["overall"],
            reverse=True
        )
    )

    return overall_pct, outcomes, unit_avgs


def compute_all_progress(ws):
    target_col = 9
    last_row = ws.max_row

    # Find last valid data row
    for r in range(ws.max_row, 0, -1):
        if is_number(ws.cell(row=r, column=target_col).value):
            last_row = r
            break

    # ------------------------------------------------------------------
    # FIRST PASS: Build all unique units excluding "ALL UNITS"
    # ------------------------------------------------------------------
    all_units = set()
    for i in range(4, last_row + 1):
        unit_cell = ws.cell(row=i, column=7).value
        if unit_cell in (None, ""):
            continue

        s = str(unit_cell).strip()
        if s.lower() == "all units":
            continue

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')

        units = [p.strip() for p in s.split(',') if p.strip()]
        all_units.update(units)

    all_unique_units = sorted(all_units)

    # Helper function
    def extract_units(unit_text):
        if not unit_text:
            return []
        s = str(unit_text).strip()
        if s.lower() == "all units":
            return all_unique_units

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')
        return [p.strip() for p in s.split(',') if p.strip()]

    # ------------------------------------------------------------------
    # SECOND PASS: Compute cumulative progress distributions
    # ------------------------------------------------------------------
    row_percentages = []
    unit_q_sums = {}
    unit_q_cnts = {}

    for i in range(4, last_row + 1):
        target_val = ws.cell(row=i, column=target_col).value

        if not is_number(target_val) or float(target_val) == 0:
            continue

        row_pcts = get_quarterly_cumulative_percentages(ws, i)
        row_percentages.append(row_pcts[3])

        unit_cell = ws.cell(row=i, column=7).value
        if unit_cell not in (None, ""):
            units = extract_units(unit_cell)
            for u in units:
                if u not in unit_q_sums:
                    unit_q_sums[u] = [0.0, 0.0, 0.0, 0.0]
                    unit_q_cnts[u] = 0

                for q_idx in range(4):
                    unit_q_sums[u][q_idx] += row_pcts[q_idx]

                unit_q_cnts[u] += 1

    # ------------------------------------------------------------------
    # Unit Averages Processing (Rounded to 2 decimal places)
    # ------------------------------------------------------------------
    unit_avgs = {}
    for u in unit_q_sums:
        cnt = unit_q_cnts[u]
        unit_avgs[u] = {
            "overall": round(unit_q_sums[u][3] / cnt, 2) if cnt > 0 else 0.0,
            "quarters": (
                [round(val / cnt, 2) for val in unit_q_sums[u]]
                if cnt > 0 else [0.0, 0.0, 0.0, 0.0]
            )
        }

    # Fetch processed outcome block configurations
    outcomes = compute_outcome_quarterly_progress(ws)

    # ------------------------------------------------------------------
    # HARMONIZED OVERALL PROGRESS: Balanced by Strategic Outcome
    # ------------------------------------------------------------------
    if outcomes:
        overall_pct = round(sum(d["overall"] for d in outcomes.values()) / len(outcomes), 2)
    else:
        overall_pct = 0.0

    # Sort implementing unit metrics in descending order
    unit_avgs = dict(
        sorted(
            unit_avgs.items(),
            key=lambda item: item[1]["overall"],
            reverse=True
        )
    )

    return overall_pct, outcomes, unit_avgs

# -------------------------------------------------------------------
# Page Decorators
# -------------------------------------------------------------------
def draw_page_border(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(0.5)
    margin = 15
    canvas.rect(margin, margin, doc.pagesize[0] - 2 * margin, doc.pagesize[1] - 2 * margin)
    canvas.restoreState()

# -------------------------------------------------------------------
# PDF Document Construction: Option 5
# -------------------------------------------------------------------
def generate_colored_report(ws):
    avg_overall, outcomes, avg_units = compute_all_progress(ws)
    if avg_overall is None or not outcomes:
        print("No valid data available to generate the report.")
        return

    gauge_path = os.path.join(tempfile.gettempdir(), "gauge_dial.png")
    draw_color_coded_gauge(avg_overall, gauge_path)
    
    reporting_period = detect_reporting_period(ws)
    
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name="OptimaHeader", fontName="Optima", fontSize=14, leading=18, alignment=1, spaceAfter=2
    )
    mini_header_style = ParagraphStyle(
        name="OptimaMiniHeader", fontName="Optima", fontSize=10, leading=14, alignment=1, spaceAfter=4
    )
    normal_style = ParagraphStyle(
        name="OptimaNormal", parent=styles["Normal"], fontName="Optima", fontSize=9, leading=13
    )
    heading_style = ParagraphStyle(
        name="OptimaHeading2", parent=styles["Heading2"], fontName="Optima", fontSize=12, spaceAfter=6, bold=True, alignment=1
    )
    
    story = []
    
    try:
        logo = Image("/Users/jud-05/Desktop/nyambane/judiciary_logo.png", width=100, height=60)
    except Exception:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)
    
    logo.hAlign = "CENTER"
    story.append(logo)
    story.append(Spacer(1, 4))
    story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", mini_header_style))
    story.append(Spacer(1, 2))
    story.append(Paragraph("<b>STAJ IMPLEMENTATION PROGRESS REPORT</b>", header_style))
    story.append(Spacer(1, 8))
    
    story.append(Paragraph(
        f"This report provides a consolidated summary of the STAJ progress status for <b>{reporting_period}</b>, "
        "highlighting overall achievement, outcome performance, and implementing unit milestones.", normal_style
    ))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph(f"<b>Overall Achievement {reporting_period}</b>", heading_style))
    overall_table = Table([[Image(gauge_path, width=180, height=110)]], colWidths=[200])
    overall_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER')
    ]))
    story.append(overall_table)
    story.append(Spacer(1, 12))
    
    story.append(Paragraph(f"<b>Outcome Progress for {reporting_period}</b>", heading_style))
    story.append(Spacer(1, 4))
    
    for k, data in outcomes.items():
        v = data["overall"]
        q_pcts = data["quarters"]
        
        story.append(Paragraph(f"<b>{k}: {v:.2f}%</b>", normal_style))
        story.append(Spacer(1, 2))
        
        bar_table = create_segmented_progress_bar(q_pcts, total_width=400.0)
        story.append(bar_table)
        story.append(Spacer(1, 10))
        
    story.append(Spacer(1, 6))
    
    story.append(Paragraph(f"<b>Lead Unit Progress during {reporting_period}</b>", heading_style))
    story.append(Spacer(1, 4))
    
    for k, data in avg_units.items():
        v = data["overall"]
        q_pcts = data["quarters"]
        
        story.append(Paragraph(f"<b>{k}: {v:.2f}%</b>", normal_style))
        story.append(Spacer(1, 2))
        
        bar_table = create_segmented_progress_bar(q_pcts, total_width=400.0)
        story.append(bar_table)
        story.append(Spacer(1, 10))
        
    save_path = macos_save_dialog(default_name="STAJ_Full_Report.pdf", prompt="Save Full PDF Report As")
    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"Report generated successfully: {save_path}")
    else:
        print("Export cancelled.")
        
    if os.path.exists(gauge_path):
        try:
            os.remove(gauge_path)
        except OSError:
            pass
        
        
def generate_unit_and_output_progress_grouped_old(ws) -> Dict:
    target_col = 9
    last_row = ws.max_row

    # Find last valid data row
    for r in range(ws.max_row, 0, -1):
        if is_number(ws.cell(row=r, column=target_col).value):
            last_row = r
            break

    # ------------------------------------------------------------------
    # FIRST PASS:
    # Build all unique units excluding "ALL UNITS"
    # ------------------------------------------------------------------
    all_units = set()

    for i in range(4, last_row + 1):
        unit_cell = ws.cell(row=i, column=7).value

        if unit_cell in (None, ""):
            continue

        s = str(unit_cell).strip()

        # Skip ALL UNITS for discovery
        if s.lower() == "all units":
            continue

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')

        units = [p.strip() for p in s.split(',') if p.strip()]
        all_units.update(units)

    all_unique_units = sorted(all_units)

    # ------------------------------------------------------------------
    # Helper function matching compute_all_progress perfectly
    # ------------------------------------------------------------------
    def extract_units(unit_text):
        if not unit_text:
            return []

        s = str(unit_text).strip()

        # Expand ALL UNITS
        if s.lower() == "all units":
            return all_unique_units

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')

        return [p.strip() for p in s.split(',') if p.strip()]

    # ------------------------------------------------------------------
    # SECOND PASS:
    # Compute progress grouped by Unit and Output
    # ------------------------------------------------------------------
    unit_progress = {}
    for row in range(4, last_row + 1):
        lead_unit_text = ws.cell(row=row, column=7).value
        output_name = ws.cell(row=row, column=4).value
        uom = ws.cell(row=row, column=8).value
        target_val = ws.cell(row=row, column=target_col).value

        if not lead_unit_text or not output_name:
            continue
        if not is_number(target_val) or float(target_val) == 0:
            continue

        # Extract values using the exact concept logic
        start_col = get_active_year_start_col(ws, target_col=9)
        q_values = []
        uom_text = str(uom).lower() if uom else ""
        is_percentage = ("percent" in uom_text) or ("%" in uom_text)

        # Re-verify target format conversion matches target processing requirements
        target_num = float(target_val)
        if is_percentage and target_num <= 1:
            target_num *= 100

        for col in range(start_col, start_col + 4):
            val = ws.cell(row=row, column=col).value
            if val is None or val == "":
                continue

            if is_percentage:
                if isinstance(val, str) and val.strip().endswith("%"):
                    try: q_values.append(float(val.strip().replace("%", "")))
                    except: pass
                elif isinstance(val, (int, float)):
                    q_values.append(val * 100 if val <= 1 else val)
            else:
                if isinstance(val, (int, float)):
                    q_values.append(float(val))
                elif isinstance(val, str):
                    try: q_values.append(float(val))
                    except: pass

        # Grab cumulative vector using the identical mechanism from compute_all_progress
        row_pcts = get_quarterly_cumulative_percentages(ws, row)
        out_pct = row_pcts[3] if row_pcts else 0.0

        # Extract units using local scope configuration definitions
        units = extract_units(lead_unit_text)
        for u in units:
            if u not in unit_progress:
                unit_progress[u] = {
                    "overall_progress": [],
                    "outputs": {}
                }

            # Handle duplicated names smoothly without overriding values
            unique_output_key = output_name
            suffix_counter = 1
            while unique_output_key in unit_progress[u]["outputs"]:
                suffix_counter += 1
                unique_output_key = f"{output_name} ({suffix_counter})"

            unit_progress[u]["outputs"][unique_output_key] = {
                "row_idx": row,
                "target": target_num,
                "achievement": sum(q_values),
                "progress": out_pct,
                "uom": uom
            }
            
            # Populate running profile matching Option 5 perfectly
            unit_progress[u]["overall_progress"].append(out_pct)

    # ------------------------------------------------------------------
    # Calculate Finalized Averages
    # ------------------------------------------------------------------
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        overall = sum(arr) / len(arr) if arr else 0
        data["overall_progress"] = min(overall, 100)

    return unit_progress


def generate_unit_and_output_progress_grouped(ws) -> dict:
    target_col = 9
    last_row = ws.max_row

    # Find last valid data row
    for r in range(ws.max_row, 0, -1):
        if is_number(ws.cell(row=r, column=target_col).value):
            last_row = r
            break

    # ------------------------------------------------------------------
    # FIRST PASS: Build all unique units excluding "ALL UNITS"
    # ------------------------------------------------------------------
    all_units = set()
    for i in range(4, last_row + 1):
        unit_cell = ws.cell(row=i, column=7).value
        if unit_cell in (None, ""):
            continue

        s = str(unit_cell).strip()
        if s.lower() == "all units":
            continue

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')

        units = [p.strip() for p in s.split(',') if p.strip()]
        all_units.update(units)

    all_unique_units = sorted(all_units)

    # Helper function
    def extract_units(unit_text):
        if not unit_text:
            return []
        s = str(unit_text).strip()
        if s.lower() == "all units":
            return all_unique_units

        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')
        return [p.strip() for p in s.split(',') if p.strip()]

    # ------------------------------------------------------------------
    # SECOND PASS: Compute progress grouped by Unit and Output
    # ------------------------------------------------------------------
    unit_progress = {}
    for row in range(4, last_row + 1):
        lead_unit_text = ws.cell(row=row, column=7).value
        output_name = ws.cell(row=row, column=4).value
        uom = ws.cell(row=row, column=8).value
        target_val = ws.cell(row=row, column=target_col).value

        if not lead_unit_text or not output_name:
            continue
        if not is_number(target_val) or float(target_val) == 0:
            continue

        # Extract values using identical concept logic
        start_col = get_active_year_start_col(ws, target_col=9)
        q_values = []
        uom_text = str(uom).lower() if uom else ""
        is_percentage = ("percent" in uom_text) or ("%" in uom_text)

        target_num = float(target_val)
        if is_percentage and target_num <= 1:
            target_num *= 100

        for col in range(start_col, start_col + 4):
            val = ws.cell(row=row, column=col).value
            if val is None or val == "":
                continue

            if is_percentage:
                if isinstance(val, str) and val.strip().endswith("%"):
                    try: q_values.append(float(val.strip().replace("%", "")))
                    except: pass
                elif isinstance(val, (int, float)):
                    q_values.append(val * 100 if val <= 1 else val)
            else:
                if isinstance(val, (int, float)):
                    q_values.append(float(val))
                elif isinstance(val, str):
                    try: q_values.append(float(val))
                    except: pass

        # Grab cumulative vector using the identical mechanism
        row_pcts = get_quarterly_cumulative_percentages(ws, row)
        out_pct = round(row_pcts[3], 2) if row_pcts else 0.0

        units = extract_units(lead_unit_text)
        for u in units:
            if u not in unit_progress:
                unit_progress[u] = {
                    "overall_progress": [],
                    "outputs": {}
                }

            unique_output_key = output_name
            suffix_counter = 1
            while unique_output_key in unit_progress[u]["outputs"]:
                suffix_counter += 1
                unique_output_key = f"{output_name} ({suffix_counter})"

            unit_progress[u]["outputs"][unique_output_key] = {
                "row_idx": row,
                "target": target_num,
                "achievement": sum(q_values),
                "progress": out_pct,
                "uom": uom
            }
            
            unit_progress[u]["overall_progress"].append(out_pct)

    # ------------------------------------------------------------------
    # Calculate Finalized Averages (Rounded to 2 decimal places)
    # ------------------------------------------------------------------
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        overall = sum(arr) / len(arr) if arr else 0
        data["overall_progress"] = min(round(overall, 2), 100.00)

    return unit_progress


def generate_pdf_unit_output_report_with_target(ws):
    data = generate_unit_and_output_progress_grouped(ws)
    if not data:
        print("No data available for Unit + Output report.")
        return

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name="Header", fontName="Optima", fontSize=12, leading=20, alignment=1
    )
    normal = ParagraphStyle(
        name="NormalOptima", parent=styles["Normal"], fontName="Optima", fontSize=8, leading=14
    )

    story = []

    try:
        logo = Image("/Users/jud-05/Desktop/nyambane/judiciary_logo.png", width=100, height=60)
    except Exception:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)
    logo.hAlign = "CENTER"

    rendered_units = 0

    for unit, info in data.items():
        if rendered_units > 0:
            story.append(PageBreak())
        rendered_units += 1

        story.append(logo)
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", header_style))
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>STAJ IMPLEMENTING UNIT PROGRESS REPORT</b>", header_style))
        story.append(Spacer(1, 10))
        story.append(
            Paragraph(
                f"<font color='blue'><b>{unit} - Overall Progress: {info['overall_progress']:.2f}%</b></font>",
                header_style
            )
        )
        story.append(Spacer(1, 5))

        sorted_outputs = sorted(
            info["outputs"].items(),
            key=lambda item: item[1].get("progress", 0),
            reverse=True
        )

        for idx, (output_name, out_data) in enumerate(sorted_outputs, start=1):
            pct = out_data.get('progress', 0)
            target_val = int(out_data.get("target", 0))
            achievement = int(out_data.get('achievement', 0))
            row_idx = out_data.get('row_idx')

            story.append(Paragraph(
                f"{idx}. {output_name}: Target {target_val}; Achievement {achievement}; Total Progress {pct:.0f}%",
                normal
            ))
            story.append(Spacer(1, 2))

            q_pcts = get_quarterly_cumulative_percentages(ws, row_idx)

            bar_table = create_segmented_progress_bar(q_pcts, total_width=400.0)
            story.append(bar_table)
            story.append(Spacer(1, 8))

        story.append(Spacer(1, 10))

    save_path = macos_save_dialog(
        default_name="Unit_Output_Progress_Report_Target.pdf",
        prompt="Save Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")



def generate_pdf_specific_units_output_report_with_target(ws):
    """
    Generates a PDF report for user-selected Implementing Unit(s) + Output Progress.
    User is prompted to enter unit name(s), e.g. OCJ or OCJ, DSPOP.
    Each selected unit starts on a new page (no blank pages).
    """

    # Prompt user to enter unit(s)
    units_input = input(
        "Enter Implementing Unit(s) separated by commas (e.g. OCJ or OCJ, DSPOP): "
    ).strip()

    if not units_input:
        print("No unit entered. Operation cancelled.")
        return

    selected_units = [u.strip().upper() for u in units_input.split(",")]

    data = generate_unit_and_output_progress_grouped(ws)
    if not data:
        print("No data available for Unit + Output report.")
        return

    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=12,
        leading=20,
        alignment=1,  # Center
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=8,
        leading=14
    )

    story = []

    # Logo
    try:
        from reportlab.platypus import Image
        logo = Image(
            "/Users/jud-05/Desktop/nyambane/judiciary_logo.png",
            width=100,
            height=60
        )
    except Exception:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)

    logo.hAlign = "CENTER"

    rendered_units = 0

    for unit, info in data.items():
        if unit.upper() not in selected_units:
            continue

        # Page break ONLY between rendered units
        if rendered_units > 0:
            story.append(PageBreak())

        rendered_units += 1

        story.append(logo)
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", header_style))
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>STAJ IMPLEMENTING UNIT PROGRESS REPORT</b>", header_style))
        story.append(Spacer(1, 10))

        story.append(
            Paragraph(
                f"<font color='blue'><b>{unit} - Overall Progress: "
                f"{info['overall_progress']:.2f}%</b></font>",
                header_style
            )
        )

        story.append(Spacer(1, 5))

        # 🔹 SORT outputs by progress (highest → lowest)
        sorted_outputs = sorted(
            info["outputs"].items(),
            key=lambda item: item[1].get("progress", 0),
            reverse=True
        )

        for idx, (output_name, out_data) in enumerate(sorted_outputs, start=1):
            pct = out_data.get("progress", 0)
            target_val = int(out_data.get("target", 0))
            achievement = int(out_data.get("achievement", 0))

            story.append(Paragraph(
                f"{idx}. {output_name}: Target {target_val}; "
                f"Achievement {achievement}; Progress {pct:.0f}%",
                normal
            ))

            total_width = 400
            bar_width = int((pct / 100) * total_width)

            # Force visibility for 0%
           #if pct == 0:
            #    bar_width = 3

            bar_color = get_bar_color(pct)

            bar = Table(
                [["", ""]],
                colWidths=[bar_width, total_width - bar_width],
                rowHeights=10
            )
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), bar_color),
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
            ]))

            story.append(bar)
            story.append(Spacer(1, 5))

        story.append(Spacer(1, 10))

    if rendered_units == 0:
        print(f"No matching units found for: {', '.join(selected_units)}")
        return

    # Dynamic file name based on selected units
    unit_suffix = "_".join(selected_units)
    file_name = f"{unit_suffix}_STAJ_Progress_Report.pdf"

    save_path = macos_save_dialog(
        default_name=file_name,
        prompt="Save Specific Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(
            story,
            onFirstPage=draw_page_border,
            onLaterPages=draw_page_border
        )
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")
        
# -------------------------------------------------------------------
# Main Menu
# -------------------------------------------------------------------
def main_menu():
    print("\n=============================================")
    print("STAJ Tools - Python Conversion (macOS-safe)")
    print("=============================================")
    print("Choose an action:")
    print("5 → CommandButton7: Combined Overall/Outcome/Lead Unit Report")
    print("8 → Full PDF Report including Unit + Output Progress")
    print("9 → Generate PDF Specific Units + Output Report with Target")
    print("---------------------------------------------")
    choice = input("Enter number (or q to quit): ").strip().lower()
    if choice == "q":
        print("Exiting application.")
        return

    if choice not in ["5", "8", "9"]:
        print("Unknown choice. Exiting.")
        return

    wb_path = macos_choose_file(prompt="Select STAJ Monitoring Tool workbook (Excel)")
    if not wb_path:
        print("No workbook selected. Exiting.")
        return

    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
    except Exception as e:
        print(f"Failed to open workbook: {e}")
        return

    target_sheet = "Working Space - Data Entry"
    if target_sheet not in wb.sheetnames:
        print(f"Workbook does not contain sheet named '{target_sheet}'. Exiting.")
        return

    ws = wb[target_sheet]

    if choice == "5":
        generate_colored_report(ws)
    elif choice == "8":
        generate_pdf_unit_output_report_with_target(ws)
    elif choice == "9":
        generate_pdf_specific_units_output_report_with_target(ws)
    print("\nDone.")

if __name__ == "__main__":
    main_menu()