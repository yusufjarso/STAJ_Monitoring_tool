# staj_dashboard.py
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import openpyxl
import matplotlib.colors as mcolors
from datetime import datetime
import re
import os
import base64
from PIL import Image
import io
import warnings
warnings.filterwarnings('ignore')

# Page configuration
st.set_page_config(
    page_title="STAJ Monitoring Dashboard",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        color: #1f3a5f;
        text-align: center;
        padding: 0.5rem 0;
        border-bottom: 3px solid #1f3a5f;
        margin-bottom: 2rem;
    }
    .kpi-card {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        text-align: center;
        margin: 0.5rem 0;
        transition: transform 0.2s;
    }
    .kpi-card:hover {
        transform: translateY(-5px);
    }
    .kpi-value {
        font-size: 2.8rem;
        font-weight: bold;
        color: #1f3a5f;
    }
    .kpi-label {
        font-size: 0.9rem;
        color: #6c757d;
        margin-top: 0.5rem;
    }
    .status-on-track {
        color: #28a745;
        font-weight: bold;
        font-size: 1.1rem;
    }
    .status-at-risk {
        color: #ffc107;
        font-weight: bold;
        font-size: 1.1rem;
    }
    .status-off-track {
        color: #dc3545;
        font-weight: bold;
        font-size: 1.1rem;
    }
    .footer {
        text-align: center;
        padding: 1rem 0;
        color: #6c757d;
        font-size: 0.85rem;
        border-top: 1px solid #e9ecef;
        margin-top: 2rem;
    }
    </style>
""", unsafe_allow_html=True)

# ============================================================================
# LOGO HANDLING
# ============================================================================

def get_logo_base64():
    """Get base64 encoded logo from image file or create a text-based logo"""
    logo_paths = [
        'image.png',
        'logo.png',
        'judiciary_logo.png',
        '/Users/jud-05/Desktop/nyambane/judiciary_logo.png'
    ]
    
    for path in logo_paths:
        if os.path.exists(path):
            try:
                with open(path, 'rb') as f:
                    img_data = f.read()
                    b64 = base64.b64encode(img_data).decode()
                    return f'data:image/png;base64,{b64}'
            except:
                continue
    
    return None

def display_logo():
    """Display the logo in the header"""
    logo_b64 = get_logo_base64()
    
    if logo_b64:
        st.markdown(f"""
        <div style="display: flex; justify-content: center; padding: 0.5rem 0;">
            <img src="{logo_b64}" style="height: 80px; width: auto;">
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="display: flex; justify-content: center; padding: 0.5rem 0;">
            <div style="background-color: #1f3a5f; border-radius: 10px; padding: 0.5rem 2rem; text-align: center;">
                <div style="color: white; font-size: 1.5rem; font-weight: bold; letter-spacing: 2px;">⚖️ JUDICIARY</div>
                <div style="color: #ffd700; font-size: 0.7rem; letter-spacing: 1px;">REPUBLIC OF KENYA</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def is_number(value) -> bool:
    """Check if value is a valid number"""
    if value in (None, "", "-", " "):
        return False
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False

def numeric_value(value):
    """Convert Excel value to numeric"""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and pd.isna(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("%", "")
    try:
        return float(text)
    except (ValueError, TypeError):
        return None

def normalise_quarter_label(value):
    """Normalise quarter labels to consistent format"""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"\s+", " ", text)
    match = re.search(
        r"FY\s*(\d{2,4})\s*/\s*(\d{2,4})\s*Q\s*([1-4])",
        text,
        flags=re.IGNORECASE
    )
    if not match:
        return None
    start_year = match.group(1)
    end_year = match.group(2)
    quarter = match.group(3)
    if len(start_year) == 4:
        start_year = start_year[-2:]
    if len(end_year) == 4:
        end_year = end_year[-2:]
    return f"FY{start_year}/{end_year} Q{quarter}"

def get_dynamic_quarter_columns(ws):
    """Dynamically detect all quarter columns"""
    quarter_columns = []
    for row in ws.iter_rows():
        for cell in row:
            label = normalise_quarter_label(cell.value)
            if label is None:
                continue
            match = re.match(r"FY(\d{2})/(\d{2}) Q([1-4])", label)
            if not match:
                continue
            fy = f"FY{match.group(1)}/{match.group(2)}"
            quarter = f"Q{match.group(3)}"
            quarter_columns.append({
                "column": cell.column,
                "label": label,
                "fy": fy,
                "quarter": quarter,
                "row": cell.row
            })
    unique = {}
    for item in quarter_columns:
        key = (item["column"], item["label"])
        if key not in unique:
            unique[key] = item
    quarter_columns = list(unique.values())
    quarter_columns.sort(key=lambda x: x["column"])
    return quarter_columns

def get_available_quarter_columns(ws):
    """Get only quarter columns that contain data"""
    all_quarters = get_dynamic_quarter_columns(ws)
    if not all_quarters:
        return []
    header_row = all_quarters[0]["row"]
    available = []
    for item in all_quarters:
        has_data = False
        for row in range(header_row + 1, ws.max_row + 1):
            val = numeric_value(ws.cell(row=row, column=item["column"]).value)
            if val is not None:
                has_data = True
                break
        if has_data:
            available.append(item)
    return available

def get_quarterly_cumulative_percentages(ws, row_idx, target_col=9):
    """Calculate cumulative progress across quarters"""
    quarter_columns = get_available_quarter_columns(ws)
    if not quarter_columns:
        return []
    target = numeric_value(ws.cell(row=row_idx, column=target_col).value)
    if target is None or target == 0:
        return []
    cumulative = 0.0
    cumulative_values = []
    for item in quarter_columns:
        val = numeric_value(ws.cell(row=row_idx, column=item["column"]).value)
        if val is not None:
            cumulative += val
        percentage = min((cumulative / target) * 100.0, 100.0)
        cumulative_values.append(round(percentage, 2))
    return cumulative_values

def extract_units(unit_text, all_unique_units):
    """Extract units from cell text"""
    if not unit_text:
        return []
    s = str(unit_text).strip()
    if s.lower() == "all units":
        return all_unique_units
    for sep in [';', '/', '&', ' and ']:
        s = s.replace(sep, ',')
    return [p.strip() for p in s.split(',') if p.strip()]

def get_bar_color(val: float) -> str:
    """Get color for progress bar"""
    v = max(0.0, min(val, 100.0))
    cmap = mcolors.LinearSegmentedColormap.from_list("rag_smooth", ["#d62728", "#FFD700", "#006400"])
    return mcolors.to_hex(cmap(v / 100.0))

def get_status_label(value):
    """Get status label based on progress"""
    if value >= 70:
        return "On Track", "status-on-track"
    elif value >= 40:
        return "At Risk", "status-at-risk"
    else:
        return "Off Track", "status-off-track"

# ============================================================================
# IMPROVED DATA PROCESSING - FIXED OUTCOME DETECTION
# ============================================================================

def compute_all_progress(ws):
    """Compute all progress metrics - improved outcome detection"""
    target_col = 9
    last_row = ws.max_row
    
    # Find last valid data row
    for r in range(ws.max_row, 0, -1):
        if is_number(ws.cell(row=r, column=target_col).value):
            last_row = r
            break
    
    # First pass: Build all unique units
    all_units = set()
    for i in range(4, last_row + 1):
        unit_cell = ws.cell(row=i, column=7).value
        if unit_cell in (None, ""):
            continue
        s = str(unit_cell).strip()
        if s.lower() == "all units":
            continue
        for sep in [';', '/', '&', ' and ']:
            s = s.replace(sep, ',')
        units = [p.strip() for p in s.split(',') if p.strip()]
        all_units.update(units)
    all_unique_units = sorted(all_units)
    
    # Second pass: Compute progress
    unit_q_sums = {}
    unit_q_cnts = {}
    outcomes = {}
    current_outcome = None
    indicator_q_pcts = []
    outcome_names = []
    
    # First, find all outcome rows by looking for "Outcome" in column A
    outcome_rows = {}
    for i in range(2, last_row + 1):
        col_a = ws.cell(row=i, column=1).value
        if col_a and isinstance(col_a, str) and "Outcome" in col_a:
            # This is an outcome row
            outcome_name = col_a.strip()
            outcome_rows[i] = outcome_name
    
    # If no "Outcome" found in column A, look in column B
    if not outcome_rows:
        for i in range(2, last_row + 1):
            col_b = ws.cell(row=i, column=2).value
            if col_b and isinstance(col_b, str) and "Outcome" in col_b:
                outcome_name = col_b.strip()
                outcome_rows[i] = outcome_name
    
    # If still no outcomes found, use the original method
    if not outcome_rows:
        # Use the original detection method
        current_outcome = 1
        for i in range(4, last_row + 1):
            target_val = ws.cell(row=i, column=target_col).value
            
            if target_val in (None, ""):
                if indicator_q_pcts:
                    q_avgs = [
                        round(sum(vals) / len(vals), 2)
                        for vals in zip(*indicator_q_pcts)
                    ]
                    outcomes[f"Outcome {current_outcome}"] = {
                        "overall": q_avgs[-1] if q_avgs else 0,
                        "quarters": q_avgs
                    }
                    current_outcome += 1
                indicator_q_pcts = []
                continue
            
            if not is_number(target_val) or float(target_val) == 0:
                continue
            
            row_pcts = get_quarterly_cumulative_percentages(ws, i)
            if row_pcts:
                indicator_q_pcts.append(row_pcts)
        
        if indicator_q_pcts:
            q_avgs = [
                round(sum(vals) / len(vals), 2)
                for vals in zip(*indicator_q_pcts)
            ]
            outcomes[f"Outcome {current_outcome}"] = {
                "overall": q_avgs[-1] if q_avgs else 0,
                "quarters": q_avgs
            }
    
    else:
        # Process each outcome using the detected outcome rows
        outcome_indices = sorted(outcome_rows.keys())
        
        for idx, start_row in enumerate(outcome_indices):
            # Determine end row (next outcome row or last row)
            if idx + 1 < len(outcome_indices):
                end_row = outcome_indices[idx + 1]
            else:
                end_row = last_row
            
            outcome_name = outcome_rows[start_row]
            
            # Collect all indicator data for this outcome
            outcome_q_pcts = []
            for row in range(start_row + 1, end_row):
                target_val = ws.cell(row=row, column=target_col).value
                if not is_number(target_val) or float(target_val) == 0:
                    continue
                
                row_pcts = get_quarterly_cumulative_percentages(ws, row)
                if row_pcts:
                    outcome_q_pcts.append(row_pcts)
                    
                    # Also track units for this outcome
                    unit_cell = ws.cell(row=row, column=7).value
                    units = extract_units(unit_cell, all_unique_units)
                    
                    for u in units:
                        if u not in unit_q_sums:
                            unit_q_sums[u] = [0.0] * len(row_pcts)
                            unit_q_cnts[u] = 0
                        for q_idx in range(len(row_pcts)):
                            unit_q_sums[u][q_idx] += row_pcts[q_idx]
                        unit_q_cnts[u] += 1
            
            if outcome_q_pcts:
                q_avgs = [
                    round(sum(vals) / len(vals), 2)
                    for vals in zip(*outcome_q_pcts)
                ]
                outcomes[outcome_name] = {
                    "overall": q_avgs[-1] if q_avgs else 0,
                    "quarters": q_avgs
                }
    
    # Calculate unit averages
    unit_avgs = {}
    for u in unit_q_sums:
        cnt = unit_q_cnts[u]
        if cnt > 0:
            unit_avgs[u] = {
                "overall": round(unit_q_sums[u][-1] / cnt, 2) if cnt > 0 else 0.0,
                "quarters": [round(val / cnt, 2) for val in unit_q_sums[u]] if cnt > 0 else [0.0]
            }
    
    # Overall progress
    if outcomes:
        overall_pct = round(sum(d["overall"] for d in outcomes.values()) / len(outcomes), 2)
    else:
        overall_pct = 0.0
    
    unit_avgs = dict(sorted(unit_avgs.items(), key=lambda item: item[1]["overall"], reverse=True))
    
    return overall_pct, outcomes, unit_avgs

def get_quarter_labels(ws):
    """Get available quarter labels"""
    quarters = get_available_quarter_columns(ws)
    return [q["label"] for q in quarters]

# ============================================================================
# LOAD DATA
# ============================================================================

@st.cache_data
def load_excel_data(file_path):
    """Load and process Excel data"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        target_sheet = "Working Space - Data Entry"
        if target_sheet not in wb.sheetnames:
            st.error(f"Sheet '{target_sheet}' not found in workbook")
            return None, None, None, None
        ws = wb[target_sheet]
        
        overall, outcomes, units = compute_all_progress(ws)
        quarter_labels = get_quarter_labels(ws)
        
        return ws, overall, outcomes, units, quarter_labels
    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        return None, None, None, None, None

# ============================================================================
# DASHBOARD VISUALIZATIONS
# ============================================================================

def create_gauge(value, title):
    """Create a gauge chart"""
    try:
        fig = go.Figure(go.Indicator(
            mode="gauge+number",
            value=value,
            title={'text': title, 'font': {'size': 16}},
            number={'suffix': "%", 'font': {'size': 28, 'color': '#1f3a5f'}},
            gauge={
                'axis': {'range': [0, 100], 'tickwidth': 1, 'tickcolor': "darkblue"},
                'bar': {'color': "darkblue", 'thickness': 0.15},
                'steps': [
                    {'range': [0, 30], 'color': "#d62728"},
                    {'range': [30, 70], 'color': "#FFD700"},
                    {'range': [70, 100], 'color': "#006400"}
                ],
                'threshold': {
                    'line': {'color': "black", 'width': 4},
                    'thickness': 0.6,
                    'value': value
                }
            }
        ))
        fig.update_layout(
            height=250,
            margin=dict(l=20, r=20, t=40, b=20),
            paper_bgcolor='rgba(0,0,0,0)'
        )
        return fig
    except Exception as e:
        return None

def create_outcome_chart(outcomes):
    """Create outcome progress chart"""
    if not outcomes:
        return None
    
    try:
        df = pd.DataFrame({
            'Outcome': list(outcomes.keys()),
            'Progress (%)': [d['overall'] for d in outcomes.values()]
        }).sort_values('Progress (%)', ascending=True)
        
        colors = [get_bar_color(v) for v in df['Progress (%)']]
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=df['Progress (%)'],
            y=df['Outcome'],
            orientation='h',
            marker_color=colors,
            text=df['Progress (%)'].apply(lambda x: f'{x:.1f}%'),
            textposition='outside',
            hovertemplate='<b>%{y}</b><br>Progress: %{x:.1f}%<extra></extra>'
        ))
        
        fig.update_layout(
            height=350,
            xaxis=dict(range=[0, 105], title="Progress (%)", gridcolor='lightgrey', gridwidth=0.5),
            yaxis=dict(title="", gridcolor='lightgrey', gridwidth=0.5),
            showlegend=False,
            margin=dict(l=10, r=10, t=20, b=20),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=12)
        )
        return fig
    except Exception as e:
        return None

def create_unit_chart(units, top_n=10):
    """Create implementing unit chart"""
    if not units:
        return None
    
    try:
        unit_items = list(units.items())
        if len(unit_items) > top_n:
            unit_items = unit_items[:top_n]
        
        df = pd.DataFrame({
            'Unit': [u[0] for u in unit_items],
            'Progress (%)': [u[1]['overall'] for u in unit_items]
        }).sort_values('Progress (%)', ascending=True)
        
        colors = [get_bar_color(v) for v in df['Progress (%)']]
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=df['Progress (%)'],
            y=df['Unit'],
            orientation='h',
            marker_color=colors,
            text=df['Progress (%)'].apply(lambda x: f'{x:.1f}%'),
            textposition='outside',
            hovertemplate='<b>%{y}</b><br>Progress: %{x:.1f}%<extra></extra>'
        ))
        
        fig.update_layout(
            height=400,
            xaxis=dict(range=[0, 105], title="Progress (%)", gridcolor='lightgrey', gridwidth=0.5),
            yaxis=dict(title="", gridcolor='lightgrey', gridwidth=0.5),
            showlegend=False,
            margin=dict(l=10, r=10, t=20, b=20),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=11)
        )
        return fig
    except Exception as e:
        return None

def create_quarterly_progress_chart(outcomes, quarter_labels):
    """Create quarterly progress trend chart"""
    if not outcomes or not quarter_labels:
        return None
    
    try:
        data = []
        for outcome, values in outcomes.items():
            q_data = values.get('quarters', [])
            q_data = q_data[:len(quarter_labels)]
            while len(q_data) < len(quarter_labels):
                q_data.append(None)
            data.append({'Outcome': outcome, 'values': q_data})
        
        fig = go.Figure()
        for item in data:
            valid_indices = [i for i, v in enumerate(item['values']) if v is not None]
            valid_labels = [quarter_labels[i] for i in valid_indices]
            valid_values = [item['values'][i] for i in valid_indices]
            
            if valid_values:
                fig.add_trace(go.Scatter(
                    x=valid_labels,
                    y=valid_values,
                    mode='lines+markers',
                    name=item['Outcome'],
                    line=dict(width=2),
                    marker=dict(size=8),
                    hovertemplate='<b>%{fullData.name}</b><br>%{x}: %{y:.1f}%<extra></extra>'
                ))
        
        fig.update_layout(
            height=350,
            xaxis=dict(title="Quarter", gridcolor='lightgrey', gridwidth=0.5),
            yaxis=dict(title="Progress (%)", range=[0, 105], gridcolor='lightgrey', gridwidth=0.5),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(l=10, r=10, t=40, b=20),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=11)
        )
        return fig
    except Exception as e:
        return None

def create_unit_heatmap(units):
    """Create heatmap of unit performance"""
    if not units:
        return None
    
    try:
        unit_items = list(units.items())
        
        max_quarters = 0
        for _, data in unit_items:
            q_len = len(data.get('quarters', []))
            if q_len > max_quarters:
                max_quarters = q_len
        
        heatmap_data = []
        unit_names = []
        for unit, data in unit_items:
            q_data = data.get('quarters', [])
            while len(q_data) < max_quarters:
                q_data.append(None)
            heatmap_data.append(q_data[:max_quarters])
            unit_names.append(unit)
        
        quarter_labels = [f"Q{i+1}" for i in range(max_quarters)]
        
        text_data = []
        for row in heatmap_data:
            text_data.append([f'{v:.1f}%' if v is not None else '-' for v in row])
        
        fig = go.Figure(data=go.Heatmap(
            z=heatmap_data,
            y=unit_names,
            x=quarter_labels,
            colorscale=[
                [0, '#d62728'], [0.3, '#d62728'],
                [0.3, '#FFD700'], [0.7, '#FFD700'],
                [0.7, '#006400'], [1, '#006400']
            ],
            zmin=0, zmax=100,
            text=text_data,
            texttemplate='%{text}',
            textfont={"size": 10},
            hovertemplate='<b>%{y}</b><br>%{x}: %{z:.1f}%<extra></extra>'
        ))
        
        fig.update_layout(
            height=max(300, len(unit_names) * 20 + 50),
            xaxis=dict(title="Quarter", gridcolor='lightgrey', gridwidth=0.5),
            yaxis=dict(title="Implementing Unit", gridcolor='lightgrey', gridwidth=0.5),
            margin=dict(l=10, r=10, t=20, b=20),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=10)
        )
        return fig
    except Exception as e:
        return None

# ============================================================================
# MAIN DASHBOARD
# ============================================================================

def main():
    # Display logo
    display_logo()
    
    # Header title
    st.markdown("""
    <div style="text-align: center; padding: 0.5rem 0 1rem 0;">
        <h1 style="color: #1f3a5f; font-size: 2rem; margin: 0;">STAJ Implementation Dashboard</h1>
        <p style="color: #6c757d; font-size: 0.9rem; margin: 0;">Strategic Transformation Agenda for Justice</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # File upload
    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded_file = st.file_uploader(
            "Upload STAJ Monitoring Tool Excel file",
            type=['xlsx', 'xls'],
            help="Upload the STAJ Monitoring Tool Excel file"
        )
    
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        use_sample = st.button("📂 Use Sample Data (if available)")
    
    file_path = None
    if uploaded_file is not None:
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(uploaded_file.getbuffer())
            file_path = tmp.name
    elif use_sample:
        possible_files = [
            'STAJ Monitoring Tool - 2025_2026 .xlsx',
            'STAJ_Monitoring_Tool_2025_2026.xlsx',
            'STAJ Monitoring Tool.xlsx'
        ]
        for f in possible_files:
            if os.path.exists(f):
                file_path = f
                break
    
    if file_path is None:
        st.info("📂 Please upload the STAJ Monitoring Tool Excel file to view the dashboard.")
        st.markdown("""
        ### Dashboard Features:
        - **Overall Implementation Status** - Gauge with color-coded progress
        - **Outcome Progress** - Bar chart showing progress by Strategic Outcome
        - **Implementing Unit Performance** - Bar chart showing unit performance
        - **Quarterly Trends** - Line chart showing progress over time
        - **Unit Heatmap** - Visual matrix of unit performance by quarter
        - **Detailed Data** - Expandable table with all metrics
        """)
        return
    
    # Load data
    with st.spinner("Loading and processing data..."):
        ws, overall, outcomes, units, quarter_labels = load_excel_data(file_path)
    
    if ws is None:
        st.error("Failed to load data. Please check the file format.")
        return
    
    # Debug: Show what outcomes were detected
    st.sidebar.header("🔍 Filters")
    

    
    # Get all outcome options
    outcome_options = list(outcomes.keys()) if outcomes else []
    
    # Select ALL outcomes by default
    default_outcomes = outcome_options.copy()
    
    selected_outcomes = st.sidebar.multiselect(
        "Select Outcomes to Display",
        outcome_options,
        default=default_outcomes
    )
    
    # Get all unit options
    unit_options = list(units.keys()) if units else []
    default_units = unit_options[:5] if len(unit_options) > 5 else unit_options
    
    selected_units = st.sidebar.multiselect(
        "Select Units to Display",
        unit_options,
        default=default_units
    )
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📊 Reporting Period")
    if quarter_labels:
        for q in quarter_labels:
            st.sidebar.markdown(f"- {q}")
    
    st.sidebar.markdown("---")
    st.sidebar.markdown(f"**Data Loaded:** {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # Filter outcomes based on selection
    filtered_outcomes = {k: v for k, v in outcomes.items() if k in selected_outcomes}
    
    # KPI Cards
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        status_label, status_class = get_status_label(overall)
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-value">{overall:.1f}%</div>
            <div class="kpi-label">Overall Implementation</div>
            <div class="{status_class}">{status_label}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        active = len([o for o in filtered_outcomes.values() if o['overall'] > 0]) if filtered_outcomes else 0
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-value">{active}</div>
            <div class="kpi-label">Active Outcomes</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        avg_outcome = np.mean([d['overall'] for d in filtered_outcomes.values()]) if filtered_outcomes else 0
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-value">{avg_outcome:.1f}%</div>
            <div class="kpi-label">Average Outcome Progress</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        total_units = len(units) if units else 0
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-value">{total_units}</div>
            <div class="kpi-label">Implementing Units</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Charts Row 1
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.subheader("📈 Overall Progress")
        gauge_fig = create_gauge(overall, "STAJ Implementation")
        if gauge_fig:
            st.plotly_chart(gauge_fig, use_container_width=True)
    
    with col2:
        st.subheader("📊 Progress by Outcome")
        outcome_fig = create_outcome_chart(filtered_outcomes)
        if outcome_fig:
            st.plotly_chart(outcome_fig, use_container_width=True)
        else:
            st.info("No outcome data available")
    
    # Charts Row 2
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🏢 Implementing Unit Performance")
        unit_fig = create_unit_chart(units, top_n=10)
        if unit_fig:
            st.plotly_chart(unit_fig, use_container_width=True)
    
    with col2:
        st.subheader("📈 Quarterly Progress Trends")
        quarter_fig = create_quarterly_progress_chart(filtered_outcomes, quarter_labels)
        if quarter_fig:
            st.plotly_chart(quarter_fig, use_container_width=True)
    
    # Heatmap
    if len(units) > 0:
        st.subheader("🔄 Unit Performance Heatmap")
        heatmap_fig = create_unit_heatmap(units)
        if heatmap_fig:
            st.plotly_chart(heatmap_fig, use_container_width=True)
    
    # Detailed Data
    with st.expander("📋 Detailed Performance Data", expanded=False):
        if outcomes:
            detailed_data = []
            for outcome, data in outcomes.items():
                detailed_data.append({
                    'Category': 'Outcome',
                    'Name': outcome,
                    'Progress (%)': data['overall'],
                    'Quarters': ', '.join([f'{q:.1f}%' for q in data.get('quarters', [])])
                })
            
            for unit, data in units.items():
                detailed_data.append({
                    'Category': 'Implementing Unit',
                    'Name': unit,
                    'Progress (%)': data['overall'],
                    'Quarters': ', '.join([f'{q:.1f}%' for q in data.get('quarters', [])])
                })
            
            df_detailed = pd.DataFrame(detailed_data).sort_values('Progress (%)', ascending=False)
            
            def color_progress(val):
                if isinstance(val, (int, float)):
                    if val >= 70:
                        return 'background-color: #d4edda; color: #155724'
                    elif val >= 40:
                        return 'background-color: #fff3cd; color: #856404'
                    else:
                        return 'background-color: #f8d7da; color: #721c24'
                return ''
            
            st.dataframe(df_detailed.style.applymap(color_progress, subset=['Progress (%)']), 
                        use_container_width=True, height=400)
    
    # Footer
    st.markdown("""
    <div class="footer">
        <p>© 2026 Judiciary of Kenya - STAJ Implementation Dashboard</p>
        <p style="font-size: 0.75rem;">Data source: STAJ Monitoring Tool - Working Space - Data Entry</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Clean up temp file
    if uploaded_file and file_path and os.path.exists(file_path):
        try:
            os.unlink(file_path)
        except:
            pass

if __name__ == "__main__":
    main()