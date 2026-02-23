import os
import subprocess
import sys
from typing import List, Tuple, Dict

import openpyxl
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet

import plotly.graph_objects as go
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Frame, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import os
# -------------------------
# macOS-native file dialogs
# -------------------------
def macos_choose_file(prompt="Select an Excel workbook"):
    script = f'''try
  set theFile to (choose file with prompt "{prompt}")
  POSIX path of theFile
on error
  return ""
end try'''
    proc = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    path = proc.stdout.strip()
    return path if path else None


def macos_save_dialog(default_name="STAJ_Report.pdf", prompt="Save PDF Report As"):
    safe_name = os.path.basename(default_name)
    script = f'''try
  set theFile to (choose file name with prompt "{prompt}" default name "{safe_name}")
  POSIX path of theFile
on error
  return ""
end try'''
    proc = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    path = proc.stdout.strip()
    return path if path else None

# -------------------------
# Utility helpers
# -------------------------
def is_number(value):
    try:
        if value is None:
            return False
        float(value)
        return True
    except Exception:
        return False


def extract_units(unit_text: str) -> List[str]:
    if not unit_text:
        return []
    s = unit_text.strip()
    for sep in [';', '/', '&', ' and ']:
        s = s.replace(sep, ',')
    parts = [p.strip() for p in s.split(',') if p.strip()]
    return parts

# -------------------------
# Gauge drawing (Plotly)
# ------------------------
import plotly.graph_objects as go
import numpy as np

def get_color_for_value(val):
    if val <= 40:
        return "red"
    elif val <= 80:
        return "gold"
    elif val <= 95:
        return "lightgreen"
    elif val <= 100:
        return "green"
    else:
        return "darkgreen"


def draw_color_coded_gauge_old(value, out_path="gauge_dial.png"):

    display_value = max(0, min(value, 100))

    domain_x = [0.0, 1.0]
    domain_y = [0.0, 0.78]

    cx = (domain_x[0] + domain_x[1]) / 2
    cy = (domain_y[0] + domain_y[1]) / 4

    angle = (1 - display_value / 100) * np.pi
    radius = 0.40

    x_tip = cx + radius * np.cos(angle)
    y_tip = cy + radius * np.sin(angle)

    fig = go.Figure()

    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=display_value,
        number={'suffix': "%", 'font': {'size': 60}, 'valueformat': '.1f'},
        domain={'x': domain_x, 'y': domain_y},
        gauge={
            'axis': {'range': [0, 100]},
            'bar': {'color': "black", 'thickness': 0.40},

            # --- Enhanced overlapping gradient steps ---
            'steps': [
                {'range': [0, 5],  'color': "#E60606"},
                {'range': [5, 10], 'color': "#D00909"},
                {'range': [10, 15], 'color': "#B30000"},
                {'range': [15, 20], 'color': "#CC0000"},
                {'range': [20, 25], 'color': "#E62E00"},
                {'range': [25, 30], 'color': "#FF4500"},
                {'range': [30, 35], 'color': "#FF6A00"},
                {'range': [35, 40], 'color': "#FF8C00"},
                {'range': [40, 45], 'color': "#FFA500"},
                {'range': [45, 50], 'color': "#FFB400"},
                {'range': [50, 55], 'color': "#FFCC00"},
                {'range': [55, 60], 'color': "#FFD700"},
                {'range': [60, 65], 'color': "#E6E600"},
                {'range': [65, 70], 'color': "#CCFF33"},
                {'range': [70, 75], 'color': "#B3FF66"},
                {'range': [75, 80], 'color': "#99FF99"},
                {'range': [80, 85], 'color': "#66FF66"},
                {'range': [85, 90], 'color': "#33FF57"},
                {'range': [90, 95], 'color': "#00E64D"},
                {'range': [95, 100], 'color': "#00CC44"},
            ]
        }
    ))

    # NEEDLE
    fig.add_shape(type="line", x0=cx, y0=cy, x1=x_tip, y1=y_tip,
                  xref='paper', yref='paper',
                  line=dict(color="black", width=4), layer='above')

    fig.add_shape(type="circle",
                  x0=cx - 0.02, y0=cy - 0.02,
                  x1=cx + 0.02, y1=cy + 0.02,
                  xref='paper', yref='paper',
                  fillcolor="black", line_color="black",
                  layer='above')

    fig.update_layout(
        width=650,
        height=600,
        paper_bgcolor="white",
        margin=dict(l=30, r=30, t=70, b=30)
    )

    fig.write_image(out_path, width=650, height=420)
    return fig



import numpy as np
import plotly.graph_objects as go

def draw_color_coded_gauge(value, out_path="gauge_dial.png"):

    display_value = max(0, min(value, 100))

    domain_x = [0.0, 1.0]
    domain_y = [0.0, 1.0]

    cx = (domain_x[0] + domain_x[1]) / 2
    cy = (domain_y[0] + domain_y[1]) / 4

    angle = (1 - display_value / 100) * np.pi
    radius = 0.40

    x_tip = cx + radius * np.cos(angle)
    y_tip = cy + radius * np.sin(angle)

    fig = go.Figure()

    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=display_value,
        number={'suffix': "%", 'font': {'size': 60}, 'valueformat': '.1f'},
        domain={'x': domain_x, 'y': domain_y},
        gauge={
            'axis': {
                'range': [0, 100],
                'tickmode': 'array',
                'tickvals': list(range(0, 101, 10)),  # Show ticks every 10%
                'ticktext': [str(i) for i in range(0, 101, 10)],
                'tickfont': {'size': 24},
            },
            'bar': {'color': "black", 'thickness': 0.40},

            'steps': [
                {'range': [0, 5],  'color': "#E60606"},
                {'range': [5, 10], 'color': "#D00909"},
                {'range': [10, 15], 'color': "#B30000"},
                {'range': [15, 20], 'color': "#CC0000"},
                {'range': [20, 25], 'color': "#E62E00"},
                {'range': [25, 30], 'color': "#FF4500"},
                {'range': [30, 35], 'color': "#FF6A00"},
                {'range': [35, 40], 'color': "#FF8C00"},
                {'range': [40, 45], 'color': "#FFA500"},
                {'range': [45, 50], 'color': "#FFB400"},
                {'range': [50, 55], 'color': "#FFCC00"},
                {'range': [55, 60], 'color': "#FFD700"},
                {'range': [60, 65], 'color': "#E6E600"},
                {'range': [65, 70], 'color': "#CCFF33"},
                {'range': [70, 75], 'color': "#B3FF66"},
                {'range': [75, 80], 'color': "#99FF99"},
                {'range': [80, 85], 'color': "#66FF66"},
                {'range': [85, 90], 'color': "#33FF57"},
                {'range': [90, 95], 'color': "#00E64D"},
                {'range': [95, 100], 'color': "#00CC44"},
            ]
        }
    ))

    # NEEDLE
    fig.add_shape(type="line", x0=cx, y0=cy, x1=x_tip, y1=y_tip,
                  xref='paper', yref='paper',
                  line=dict(color="black", width=4), layer='above')

    fig.add_shape(type="circle",
                  x0=cx - 0.02, y0=cy - 0.02,
                  x1=cx + 0.02, y1=cy + 0.02,
                  xref='paper', yref='paper',
                  fillcolor="black", line_color="black",
                  layer='above')

    fig.update_layout(
        width=650,
        height=600,
        paper_bgcolor="white",
        margin=dict(l=30, r=30, t=70, b=30)
    )

    fig.write_image(out_path, width=650, height=420)
    return fig


# -------------------------
#Detect Last Filed Quarter
#-----------------------

def detect_reporting_period(ws, target_col=8, qcols=list(range(9, 21))):
    """
    Detects the latest filled quarter across all outcome rows.
    Only considers rows with a target value (ignores blank separator rows).
    Returns string like: 'Quarter 4 of FY 2025/2026'
    """

    # Map worksheet columns to FY and Quarter
    col_map = {
        9:  ("FY 2024/2025", 1),
        10: ("FY 2024/2025", 2),
        11: ("FY 2024/2025", 3),
        12: ("FY 2024/2025", 4),
        13: ("FY 2025/2026", 1),
        14: ("FY 2025/2026", 2),
        15: ("FY 2025/2026", 3),
        16: ("FY 2025/2026", 4),
        17: ("FY 2026/2027", 1),
        18: ("FY 2026/2027", 2),
        19: ("FY 2026/2027", 3),
        20: ("FY 2026/2027", 4),
    }

    last_filled = None
    last_col_index = -1

    # Start from row 4 where outcome rows begin
    for row in range(4, ws.max_row + 1):
        target_value = ws.cell(row=row, column=target_col).value
        if target_value in (None, ""):
            continue  # skip blank separator rows

        for col, (fy, qnum) in col_map.items():
            val = ws.cell(row=row, column=col).value
            if val not in (None, "", "-", " "):
                if col > last_col_index:  # pick rightmost filled quarter
                    last_col_index = col
                    last_filled = (fy, qnum)

    if last_filled:
        fy, qnum = last_filled
        return f"Quarter {qnum} of {fy}"
    else:
        return "Reporting Period Undetermined"


def get_last_filled_quarter_for_sheet(row, ws):
    """
    Returns the last filled quarter for a given row
    """
    col_map = {
        9:  ("FY 2024/2025", 1),
        10: ("FY 2024/2025", 2),
        11: ("FY 2024/2025", 3),
        13: ("FY 2025/2026", 1),
        14: ("FY 2025/2026", 2),
        15: ("FY 2025/2026", 3),
        16: ("FY 2025/2026", 4),
        17: ("FY 2026/2027", 1),
        18: ("FY 2026/2027", 2),
        19: ("FY 2026/2027", 3),
        20: ("FY 2026/2027", 4),
    }

    last_filled = None
    for col in col_map.keys():
        val = ws.cell(row=row, column=col).value
        if val not in (None, "", "-", " "):
            last_filled = col_map[col]

    if last_filled:
        fy, qnum = last_filled
        return f"Quarter {qnum} of {fy}"
    else:
        return None




# -------------------------
# PDF generation
# -------------------------
def build_simple_pdf(title_text: str, lines: List[str], gauge_image_path: str, out_pdf_path: str):
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f"<b>{title_text}</b>", styles["Title"]))
    story.append(Spacer(1, 12))
    
       # --- 4. Set Styles ---
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'titleStyle',
        parent=styles['Title'],
        fontSize=20,
        alignment=1,  # center
        spaceAfter=20,
        textColor=colors.HexColor("#003366")
    )
    normal_style = ParagraphStyle(
        'normalStyle',
        parent=styles['Normal'],
        fontSize=12,
        leading=16
    )
    
    for ln in lines:
        story.append(Paragraph(ln.replace("&", "&amp;"), styles["Normal"]))
        story.append(Spacer(1, 6))
    if gauge_image_path and os.path.exists(gauge_image_path):
        story.append(Spacer(1, 12))
        story.append(Image(gauge_image_path, width=420, height=280))
    doc = SimpleDocTemplate(out_pdf_path, pagesize=A4)
    doc.build(story)


# -------------------------
# Core calculations
# -------------------------


def compute_outcome_progress(ws) -> Dict[str, float]:
    target_col = 8
    qcols = list(range(9, 21))   # Q1–Q4 columns
    last_row = ws.max_row

    current_outcome = 1
    outcomes = {}

    indicator_pcts = []

    for i in range(2, last_row + 1):
        target_value = ws.cell(row=i, column=target_col).value

        # --- Outcome separator (blank target row) ---
        if target_value in (None, ""):
            if len(indicator_pcts) > 0:
                outcomes[f"Outcome {current_outcome}"] = sum(indicator_pcts) / len(indicator_pcts)
                current_outcome += 1
            indicator_pcts = []
            continue

        # --- Skip non-numeric target rows ---
        if not is_number(target_value):
            continue

        target = float(target_value)

        # --- Compute achieved ---
        achieved = 0.0
        for c in qcols:
            v = ws.cell(row=i, column=c).value
            if is_number(v):
                achieved += float(v)

        # --- Compute percentage (avoid division by zero) ---
        if target > 0:
            pct = (achieved / target) * 100
        else:
            pct = 0

        # --- CAP each individual indicator at 100% ---
        pct = min(pct, 100.0)

        indicator_pcts.append(pct)

    # --- Handle last outcome ---
    if len(indicator_pcts) > 0:
        outcomes[f"Outcome {current_outcome}"] = sum(indicator_pcts) / len(indicator_pcts)

    return outcomes



def compute_lead_unit_cumulative(ws) -> Dict[str, float]:
    last_row_lead_units = 0
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=6).value not in (None, ""):
            last_row_lead_units = r
            break
    if last_row_lead_units == 0:
        return {}
    dict_sum = {}
    dict_count = {}
    for i in range(2, last_row_lead_units + 1):
        unit_cell = ws.cell(row=i, column=6).value
        if unit_cell in (None, ""):
            continue
        units = extract_units(str(unit_cell))
        target_value = ws.cell(row=i, column=8).value
        if not is_number(target_value) or float(target_value) == 0:
            continue
        target_value = float(target_value)
        totalNumber = 0.0
        for c in range(9, 21):
            v = ws.cell(row=i, column=c).value
            if is_number(v):
                totalNumber += float(v)
        progress_pct = (totalNumber / target_value) * 100 if target_value != 0 else 0.0
        for u in units:
            dict_sum[u] = dict_sum.get(u, 0.0) + progress_pct
            dict_count[u] = dict_count.get(u, 0) + 1
    avg_dict = {}
    for k, s in dict_sum.items():
        avg_dict[k] = s / dict_count[k] if dict_count[k] != 0 else 0.0
    return avg_dict


def compute_overall_pct_from_quarters(ws) -> Tuple[float, float, float]:
    target_col = 8
    qcols = list(range(9, 21))
    last_row = ws.max_row

    # Find last valid row with a target
    for r in range(ws.max_row, 0, -1):
        if is_number(ws.cell(row=r, column=target_col).value):
            last_row = r
            break

    indicator_pcts = []
    totalAchievement = 0.0
    totalTarget = 0.0

    for i in range(4, last_row + 1):
        target_val = ws.cell(row=i, column=target_col).value
        if not is_number(target_val) or float(target_val) == 0:
            continue

        target_val = float(target_val)
        totalTarget += target_val

        # Sum quarterly achievements
        ach = 0.0
        for c in qcols:
            v = ws.cell(row=i, column=c).value
            if is_number(v):
                ach += float(v)

        totalAchievement += ach

        # Compute raw percentage
        row_pct = ach / target_val

        # --- CAP AT 100% BEFORE AVERAGING ---
        row_pct = min(row_pct, 1.0)

        indicator_pcts.append(row_pct)

    # Average of capped percentages * 100
    overall_pct = (sum(indicator_pcts) / len(indicator_pcts)) * 100 if indicator_pcts else 0.0

    return overall_pct, totalAchievement, totalTarget


def compute_all_progress(ws):
    """
    Computes:
    1. Overall achievement percentage (0-100)
    2. Outcome-level progress (via compute_outcome_progress)
    3. Lead-unit averages (0-100)

    This function fully replaces:
       - report_generator_all
       - compute_overall_pct_from_quarters
    """

    target_col = 8
    qcols = list(range(9, 21))  # Q1–Q4 columns
    last_row = ws.max_row

    # --- Identify the last row containing a numeric target ---
    for r in range(ws.max_row, 0, -1):
        if is_number(ws.cell(row=r, column=target_col).value):
            last_row = r
            break

    # Storage
    row_percentages = []   # per-indicator capped percentages
    unit_sum = {}          # lead unit performance sum
    unit_cnt = {}          # lead unit entry count

    # --- Loop over indicators ---
    for i in range(4, last_row + 1):
        target_val = ws.cell(row=i, column=target_col).value

        if not is_number(target_val) or float(target_val) == 0:
            continue

        target_val = float(target_val)

        # --- Sum quarterly achievements ---
        ach = 0.0
        for c in qcols:
            v = ws.cell(row=i, column=c).value
            if is_number(v):
                ach += float(v)

        # --- Compute capped row-level percentage (0–100) ---
        row_frac = ach / target_val
        row_frac = min(row_frac, 1.0)
        row_pct = row_frac * 100

        row_percentages.append(row_pct)

        # --- Lead Unit Aggregation ---
        unit_cell = ws.cell(row=i, column=6).value
        if unit_cell not in (None, ""):
            units = extract_units(str(unit_cell))
            for u in units:
                unit_sum[u] = unit_sum.get(u, 0.0) + row_pct
                unit_cnt[u] = unit_cnt.get(u, 0) + 1

    # --- Overall percentage ---
    overall_pct = sum(row_percentages) / len(row_percentages) if row_percentages else 0.0

    # --- Lead-unit averages ---
    unit_avgs = {u: unit_sum[u] / unit_cnt[u] for u in unit_sum}

    # --- Outcome-level progress ---
    outcomes = compute_outcome_progress(ws)

    #sort unit_avgs by percentage descending
    unit_avgs = dict(sorted(unit_avgs.items(), key=lambda item: item[1], reverse=True))
    
    return overall_pct, outcomes, unit_avgs


    build_simple_pdf(title, lines, gauge_path, save_path)
    print(f"PDF generated and saved at: {save_path}")
    print(f"PDF generated and saved at: {save_path}")
    # --- 5. Cleanup Gauge Image ---
    if os.path.exists(gauge_path):
        os.remove(gauge_path)


        
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
'''
# --- Add this helper ---
def get_bar_color(val):
    """Match gauge colors for progress bars"""
    if val <= 50:
        return colors.red
    elif val <= 75:
        return colors.orange  # amber
    elif val <= 94:
        return colors.green
    elif val <= 100:
        return colors.darkgreen
    else:
        return colors.darkgreen

'''

from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

# macOS standard Optima font paths
try:
    pdfmetrics.registerFont(TTFont('Optima', '/System/Library/Fonts/Optima.ttc'))
    pdfmetrics.registerFont(TTFont('Optima-Bold', '/System/Library/Fonts/Optima.ttc'))
except:
    print("Warning: Optima font not found; using default fonts.")


def get_bar_color(val):
    """
    Returns a color matching the smooth gauge transition:
    darkred → red → orangered → gold → lightgreen → green
    """

    # Clamp 0–100
    v = max(0, min(val, 100))

    if v <= 15:
        return "#E43434"      # darkred
    elif v <= 30:
        return "#891919"      # red
    elif v <= 50:
        return "#421708"      # orangered
    elif v <= 75:
        return "#FFD700"      # gold/amber
    elif v <= 94:
        return "#90EE90"      # lightgreen
    else:
        return "#008000"      # strong green

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
)
from reportlab.platypus import PageBreak
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch

import os

# -------------------------------------------------------------------
# PAGE BORDER DRAWER
# -------------------------------------------------------------------
def draw_page_border_old(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(2)

    # Create a clean outer border
    margin = 5.0
    canvas.rect(
        margin,
        margin,
        doc.pagesize[0] - 2 * margin,
        doc.pagesize[1] - 2 * margin
    )
    canvas.restoreState()

# -------------------------------------------------------------------
# PAGE BORDER DRAWER
# -------------------------------------------------------------------
def draw_page_border(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(2)

    # Create a clean outer border
    margin = 20
    canvas.rect(
        margin,
        margin,
        doc.pagesize[0] - 2 * margin,
        doc.pagesize[1] - 2 * margin
    )
    canvas.restoreState()


# -------------------------------------------------------------------
# MAIN FUNCTION
# -------------------------------------------------------------------
def commandbutton7_colored_reportxx(ws):
    """
    Generates a Judiciary-styled full PDF report with:
    - Judiciary logo
    - Borders
    - Gauge image
    - Outcome & lead-unit progress bars
    """
    avgOverall, outcomes, avg_units = compute_all_progress(ws)

    if avgOverall is None:
        print("No data available to generate report.")
        return

    # Draw gauge
    gauge_path = "gauge_dial.png"
    draw_color_coded_gauge(avgOverall, gauge_path)

    # --- Styles ---
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    title_style = styles["Title"]
    heading = styles["Heading2"]

    # --- Build Story ---
    story = []

    # --------------------------------------------------------------
    # HEADER WITH LOGO + TITLE
    # --------------------------------------------------------------
    try:
        logo = Image("/Users/jud-05/Desktop/nyambane/judiciary_logo2.png", width=60, height=60)
    except:
        logo = Paragraph("<b>JUDICIARY</b>", normal)

    header_table = Table(
        [
            [logo,
             Paragraph("<b>JUDICIARY OF KENYA<br/>STAJ<br/> IMPLEMENTATION REPORT</b>",
                       title_style)
            ]
        ],
        colWidths=[90, 200]
    )

    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ]))

    story.append(header_table)
    story.append(Spacer(1, 20))

    # --------------------------------------------------------------
    # INTRODUCTION
    # --------------------------------------------------------------
    story.append(Paragraph(
        "This report provides a consolidated summary of the current "
        "status of STAJ implementation, based on progress submissions "
        "from implementing units across the Judiciary.", normal
    ))
    story.append(Spacer(1, 10))

    story.append(Paragraph(
        "It highlights overall achievement, outcome-level performance, "
        "and lead-unit progress to support evidence-based decision-making.",
        normal
    ))
    story.append(Spacer(1, 20))

    # --------------------------------------------------------------
    # OVERALL PROGRESS + GAUGE
    # --------------------------------------------------------------
    story.append(Paragraph("<b>Overall Achievement</b>", heading))

    overall_data = [
        [
            Paragraph(f"<b>{avgOverall:.1f}%</b>", normal),
            Image(gauge_path, width=200, height=120)
        ]
    ]
    overall_table = Table(overall_data, colWidths=[200, 300])
    overall_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER')
    ]))

    story.append(overall_table)
    story.append(Spacer(1, 20))

    # --------------------------------------------------------------
    # OUTCOME PROGRESS SECTION
    # --------------------------------------------------------------
# -----------------------------------------------------------
# OUTCOME PROGRESS (UPDATED WITH PERIOD)
# -----------------------------------------------------------

    reporting_period = detect_reporting_period(ws)

    story.append(Paragraph(
        f"<b>Outcome Progress for {reporting_period}</b>",
        normal
    ))
    story.append(Spacer(1, 6))

    for k, v in outcomes.items():
        story.append(Paragraph(f"{k}: {v:.1f}%", normal))

        bar_width = int((v / 100) * 400)
        bar_color = get_bar_color(v)

        bar = Table(
        [["", ""]],
        colWidths=[bar_width, 400 - bar_width],
        rowHeights=10
    )
    bar.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,0), bar_color),
        ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
    ]))

    story.append(bar)
    story.append(Spacer(1, 6))

    story.append(Spacer(1, 15))

    # --------------------------------------------------------------
    # LEAD UNIT PROGRESS SECTION
    # --------------------------------------------------------------
    story.append(Paragraph("<b>Lead Unit Progress</b>", heading))
    story.append(Spacer(1, 10))

    for k, v in avg_units.items():
        story.append(Paragraph(f"{k}: {v:.1f}%", normal))

        bar_width = int((v / 100) * 400)
        bar_color = get_bar_color(v)

        bar = Table(
            [["", ""]],
            colWidths=[bar_width, 400 - bar_width],
            rowHeights=10
        )
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))

        story.append(bar)
        story.append(Spacer(1, 6))

    # --------------------------------------------------------------
    # SAVE PDF
    # --------------------------------------------------------------
    save_path = macos_save_dialog(
        default_name="STAJ_Full_Report.pdf",
        prompt="Save Full PDF Report As"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"Full report saved as: {save_path}")
    else:
        print("Save cancelled.")

    # Cleanup gauge image
    try:
        os.remove(gauge_path)
    except:
        pass
    

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import os

# Register Optima font (ensure Optima.ttf is available OR use system font name)
try:
    pdfmetrics.registerFont(TTFont('Optima', 'Optima.ttf'))
except:
    pass  # System-installed Optima will be used if available


# -------------------------------------------------------------------
# REDUCED BORDER (thin)
# -------------------------------------------------------------------
def draw_page_border(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(0.5)   # Reduced from 2 → 1

    margin = 15  # Smaller border
    canvas.rect(
        margin,
        margin,
        doc.pagesize[0] - 2 * margin,
        doc.pagesize[1] - 2 * margin
    )
    canvas.restoreState()

# -------------------------------------------------------
# QUARTERLY EXTRACTOR (12 columns → 4 quarter averages)
# -------------------------------------------------------
def compute_per_fy_quarter_progress(ws, row_index):
    """
    Reads quarterly values stored in 12 columns covering 3 fiscal years.
    Assumed layout:
    Col 2-13 (B–M) = Q1Y1..Q4Y3
    """

    # Quarter mapping (3 years per quarter)
    qmap = {
        "Q1": [2, 6, 10],  # B, F, J
        "Q2": [3, 7, 11],  # C, G, K
        "Q3": [4, 8, 12],  # D, H, L
        "Q4": [5, 9, 13],  # E, I, M
    }

    out = {}

    for q, cols in qmap.items():
        vals = []
        for c in cols:
            v = ws.cell(row=row_index, column=c).value
            try:
                v = float(v)
                vals.append(v)
            except:
                pass

        out[q] = sum(vals) / len(vals) if vals else 0.0

    return out


# -------------------------------------------------------
# COLOUR SEQUENCE FOR QUARTERLY PROGRESSION
# -------------------------------------------------------
def quarter_color(q):
    if q == "Q1":
        return colors.red
    if q == "Q2":
        return colors.Color(1, 0.4, 0.4)  # light red
    if q == "Q3":
        return colors.orange
    if q == "Q4":
        return colors.green  # final quarter - OK colour
    return colors.grey


# -------------------------------------------------------
# STACKED QUARTER BAR (4 segments)
# -------------------------------------------------------
def draw_quarterly_progress_bar(qvalues):
    """
    qvalues = {"Q1": xx, "Q2": yy, "Q3": zz, "Q4": vv}
    Creates a stacked bar with colours per quarter.
    """

    total_width = 400
    segments = []
    styles = []

    for i, q in enumerate(["Q1", "Q2", "Q3", "Q4"]):
        colour = quarter_color(q)
        segments.append("")  # dummy cell
        styles.append(('BACKGROUND', (i, 0), (i, 0), colour))

    table = Table([segments], colWidths=[total_width/4]*4, rowHeights=14)

    table.setStyle(TableStyle(styles + [
        ('BOX', (0,0), (-1,-1), 0.4, colors.grey),
        ('INNERGRID', (0,0), (-1,-1), 0.3, colors.white),
    ]))

    return table



# ===================================================================
# FULL UPDATED REPORT FUNCTION WITH QUARTERLY BARS INTEGRATED
# ===================================================================
def commandbutton7_colored_reportYxY(ws):
    """
    Quarterly-augmented full PDF report
    """

    avgOverall, outcomes, avg_units = compute_all_progress(ws)

    if avgOverall is None:
        print("No data available to generate report.")
        return

    gauge_path = "gauge_dial.png"
    draw_color_coded_gauge(avgOverall, gauge_path)

    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=14,
        leading=20,
        alignment=1,
    )
    
    mini_header_style = ParagraphStyle(
        name="MiniHeader",
        fontName="Optima",
        fontSize=10,
        leading=14,
        alignment=1,
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=9,
        leading=14
    )
    
    heading = ParagraphStyle(
        name="Heading2",
        parent=styles["Heading2"],
        fontName="Optima",
        fontSize=12,
        spaceAfter=2,
        bold=True,
        alignment=1,
    )

    story = []

    # LOGO
    try:
        logo = Image("/Users/jud-05/Desktop/nyambane/judiciary_logo.png", width=100, height=60)
    except:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)

    logo.hAlign = "CENTER"
    story.append(logo)
    story.append(Spacer(1, 5))

    story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", mini_header_style))
    story.append(Spacer(1, 5))

    story.append(Paragraph("<b>STAJ IMPLEMENTATION FULL REPORT</b>", header_style))
    story.append(Spacer(1, 10))

    story.append(Paragraph(
        "This report provides a consolidated summary of STAJ implementation progress "
        "based on submissions from implementing units across the Judiciary.",
        normal
    ))
    story.append(Spacer(1, 5))

    # OVERALL ACHIEVEMENT
    story.append(Paragraph("<b>Overall Achievement</b>", heading))

    overall_data = [
        [Image(gauge_path, width=200, height=120)]
    ]
    overall_table = Table(overall_data, colWidths=[200])
    overall_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER')
    ]))
    story.append(overall_table)
    story.append(Spacer(1, 10))

    # OUTCOME PROGRESS + QUARTER BARS
    reporting_period = detect_reporting_period(ws)
    story.append(Paragraph(f"<b>Outcome Progress for {reporting_period}</b>", heading))
    story.append(Spacer(1, 6))

    for r, (k, v) in enumerate(outcomes.items(), start=2):   # row index assumed matching
        story.append(Paragraph(f"<b>{k}: {v:.2f}%</b>", normal))

        # Quarterly breakdown
        qvals = compute_per_fy_quarter_progress(ws, r)
        qbar = draw_quarterly_progress_bar(qvals)
        story.append(qbar)

        qtext = " | ".join([f"{q}: {qvals[q]:.1f}%" for q in ["Q1","Q2","Q3","Q4"]])
        story.append(Paragraph(qtext, normal))

        story.append(Spacer(1, 8))

    # IMPLEMENTING UNIT PROGRESS
    story.append(Paragraph("<b>Implementing Unit Progress</b>", heading))
    story.append(Spacer(1, 6))

    for r, (k, v) in enumerate(avg_units.items(), start=20):  # different row start
        story.append(Paragraph(f"<b>{k}: {v:.2f}%</b>", normal))

        qvals = compute_per_fy_quarter_progress(ws, r)
        qbar = draw_quarterly_progress_bar(qvals)
        story.append(qbar)

        qtext = " | ".join([f"{q}: {qvals[q]:.1f}%" for q in ["Q1","Q2","Q3","Q4"]])
        story.append(Paragraph(qtext, normal))

        story.append(Spacer(1, 8))

    # SAVE
    save_path = macos_save_dialog(
        default_name="STAJ_Full_Report.pdf",
        prompt="Save Full PDF Report As"
    )

    if save_path:
        doc = SimpleDocTemplate(
            save_path,
            pagesize=A4,
            leftMargin=25,
            rightMargin=25,
            topMargin=25,
            bottomMargin=25
        )

        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print("Report saved:", save_path)

    try:
        os.remove(gauge_path)
    except:
        pass

# -------------------------------------------------------------------
# MAIN REPORT FUNCTION
# -------------------------------------------------------------------
def commandbutton7_colored_reportYY(ws):
    """
    Judiciary-styled full PDF report:
    - Centered logo
    - Reduced header font
    - Optima font
    - Thinner borders
    """
    avgOverall, outcomes, avg_units = compute_all_progress(ws)

    if avgOverall is None:
        print("No data available to generate report.")
        return

    # Generate gauge image
    gauge_path = "gauge_dial.png"
    draw_color_coded_gauge(avgOverall, gauge_path)

    # -----------------------------------------------------------
    # Styles
    # -----------------------------------------------------------
    styles = getSampleStyleSheet()

    # Header in Optima, smaller font
    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=14,       # Reduced from 22 → 16
        leading=20,
        alignment=1,       # Centered
        bold=True,
    )
    
    mini_header_style = ParagraphStyle(
        name="MiniHeader",
        fontName="Optima",
        fontSize=10,
        leading=14,
        alignment=1,
        bold=True,
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=9,
        leading=14,
        bold=True
    )
    
    heading = ParagraphStyle(
        name="Heading2",
        parent=styles["Heading2"],
        fontName="Optima",
        fontSize=12,
        spaceAfter=2,
        bold=True,
        alignment=1,  # Centered
    )

    # -----------------------------------------------------------
    # PDF Story
    # -----------------------------------------------------------
    story = []

    # -----------------------------------------------------------
    # CENTERED LOGO
    # -----------------------------------------------------------
    try:
        logo = Image("/Users/jud-05/Desktop/nyambane/judiciary_logo.png", width=100, height=60)
    except:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)

    logo.hAlign = "CENTER"
    story.append(logo)
    story.append(Spacer(1, 5))

    # -----------------------------------------------------------
    # HEADER TITLE (Optima font)
    # -----------------------------------------------------------
    story.append(Paragraph(
        "<b>JUDICIARY OF KENYA</b>",
        mini_header_style
    ))
    story.append(Spacer(1, 5))

    # HEADER TITLE (Optima font)
    # -----------------------------------------------------------
    story.append(Paragraph(
        "<b>STAJ IMPLEMENTATION FULL REPORT</b>",
        header_style
    ))
    story.append(Spacer(1, 10))
    
    # -----------------------------------------------------------
    # INTRODUCTION
    # -----------------------------------------------------------
    story.append(Paragraph(
        "This report provides a consolidated summary of the current "
        "status of STAJ implementation, based on progress submissions "
        "from implementing units across the Judiciary. "
        "It highlights overall achievement, outcome-level performance, "
        "and Implementing-Unit progress to support evidence-based decision-making.", normal
    ))
    story.append(Spacer(1, 5))

    #story.append(Paragraph(
     #   "It highlights overall achievement, outcome-level performance, "
      #  "and Implementing-Unit progress to support evidence-based decision-making.",
       # normal
   # ))
    #story.append(Spacer(1, 5))

    # -----------------------------------------------------------
    # OVERALL ACHIEVEMENT WITH GAUGE
    # -----------------------------------------------------------
    #story.append(Paragraph(f"<b>Overall Achievement: {avgOverall:.1f}%</b>", heading))
    story.append(Paragraph("<b>Overall Achievement</b>", heading))
   
    overall_data = [
        [
            Image(gauge_path, width=200, height=120)
        ]
    ]
    overall_table = Table(overall_data, colWidths=[200, 300])
    overall_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER')
    ]))

    story.append(overall_table)
    story.append(Spacer(1, 5))
        
    # -----------------------------------------------------------
    # OUTCOME PROGRESS
    # -----------------------------------------------------------
    reporting_period = detect_reporting_period(ws)
    story.append(Paragraph(
    f"<b>Outcome Progress for {reporting_period}</b>",
    heading
    ))

    story.append(Spacer(1, 6))
    
    for k, v in outcomes.items():
        story.append(Paragraph(f"{k}: {v:.2f}%", normal))

        bar_width = int((v / 100) * 400)
        bar_color = get_bar_color(v)

        bar = Table(
            [["", ""]],
            colWidths=[bar_width, 400 - bar_width],
            rowHeights=10
        )
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))

        story.append(bar)
        story.append(Spacer(1, 6))

    story.append(Spacer(1, 5))

    # -----------------------------------------------------------
    # LEAD UNIT PROGRESS
    # -----------------------------------------------------------
    story.append(Paragraph("<b>Implementing Unit Progress</b>", heading))
    story.append(Spacer(1, 6))

    for k, v in avg_units.items(): 
        story.append(Paragraph(f"{k}: {v:.2f}%", normal))

        bar_width = int((v / 100) * 400)
        bar_color = get_bar_color(v)

        bar = Table(
            [["", ""]],
            colWidths=[bar_width, 400 - bar_width],
            rowHeights=10
        )
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))

        story.append(bar)
        story.append(Spacer(1, 6))

    # -----------------------------------------------------------
    # SAVE FILE
    # -----------------------------------------------------------
    save_path = macos_save_dialog(
        default_name="STAJ_Full_Report.pdf",
        prompt="Save Full PDF Report As"
    )

    if save_path:
        doc = SimpleDocTemplate(
        save_path,
        pagesize=A4,
        leftMargin=25,
        rightMargin=25,
        topMargin=25,     # REDUCED from 72 → 25
        bottomMargin=25
)

        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"Full report saved at: {save_path}")
    else:
        print("Save cancelled.")

    # Clean up gauge image
    try:
        os.remove(gauge_path)
    except:
        pass


def commandbutton7_colored_report(ws):
    """
    Judiciary-styled full PDF report:
    - Centered logo
    - Quarter/FY auto-detection
    - Optima font
    - Thin border
    """
    avgOverall, outcomes, avg_units = compute_all_progress(ws)

    if avgOverall is None:
        print("No data available to generate report.")
        return

    # Generate gauge image
    gauge_path = "gauge_dial.png"
    draw_color_coded_gauge(avgOverall, gauge_path)

    # -----------------------------------------------------------
    # Styles
    # -----------------------------------------------------------
    styles = getSampleStyleSheet()

    # Header in Optima, smaller font
    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=16,       # Reduced from 22 → 16
        leading=20,
        alignment=1,       # Centered
    )
    
    mini_header_style = ParagraphStyle(
        name="MiniHeader",
        fontName="Optima",
        fontSize=12,
        leading=14,
        alignment=1,
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=11,
        leading=14
    )
    
    heading = ParagraphStyle(
        name="Heading2",
        parent=styles["Heading2"],
        fontName="Optima",
        fontSize=13,
        spaceAfter=6
    )

    # -----------------------------------------------------------
    # PDF Story
    # -----------------------------------------------------------
    story = []

    # -----------------------------------------------------------
    # CENTERED LOGO
    # -----------------------------------------------------------
    try:
        logo = Image("/Users/jud-05/Desktop/nyambane/judiciary_logo.png", width=100, height=60)
    except:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)

    logo.hAlign = "CENTER"
    story.append(logo)
    story.append(Spacer(1, 10))

    # -----------------------------------------------------------
    # HEADER TITLE (Optima font)
    # -----------------------------------------------------------
    story.append(Paragraph(
        "<b>JUDICIARY OF KENYA</b>",
        mini_header_style
    ))
    story.append(Spacer(1, 5))

    # HEADER TITLE (Optima font)
    # -----------------------------------------------------------
    story.append(Paragraph(
        "<b>STAJ IMPLEMENTATION FULL REPORT</b>",
        header_style
    ))
    story.append(Spacer(1, 10))
    
    # -----------------------------------------------------------
    # INTRODUCTION
    # -----------------------------------------------------------
    story.append(Paragraph(
        "This report provides a consolidated summary of the current "
        "status of STAJ implementation, based on progress submissions "
        "from implementing units across the Judiciary. "
        "It highlights overall achievement, outcome-level performance, "
        "and Implementing-Unit progress to support evidence-based decision-making.", normal
    ))
    story.append(Spacer(1, 5))

    #story.append(Paragraph(
     #   "It highlights overall achievement, outcome-level performance, "
      #  "and Implementing-Unit progress to support evidence-based decision-making.",
       # normal
   # ))
    story.append(Spacer(1, 10))

    # -----------------------------------------------------------
    # OVERALL ACHIEVEMENT WITH GAUGE
    # -----------------------------------------------------------
    #story.append(Paragraph(f"<b>Overall Achievement: {avgOverall:.1f}%</b>", heading))
    story.append(Paragraph("<b>Overall Achievement</b>", heading))
   
    overall_data = [
        [
            Image(gauge_path, width=200, height=120)
        ]
    ]
    overall_table = Table(overall_data, colWidths=[200, 300])
    overall_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER')
    ]))

    story.append(overall_table)
    story.append(Spacer(1, 20))
     
    # -----------------------------------------------------------
    # OUTCOME PROGRESS SECTION with PERIOD AUTO-INSERTED
    # -----------------------------------------------------------
    story.append(Paragraph(f"<b>Outcome Progress for {detect_reporting_period}</b>", heading))
    story.append(Spacer(1, 8))

    for k, v in outcomes.items():
        story.append(Paragraph(f"{k}: {v:.1f}%", normal))

        bar_width = int((v / 100) * 400)
        bar_color = get_bar_color(v)

        bar = Table([["", ""]],
                    colWidths=[bar_width, 400 - bar_width],
                    rowHeights=10)

        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))

        story.append(bar)
        story.append(Spacer(1, 6))

    story.append(Spacer(1, 15))

    # -----------------------------------------------------------
    # IMPLEMENTING UNIT PROGRESS
    # -----------------------------------------------------------
    story.append(Paragraph("<b>Implementing Unit Progress</b>", heading))
    story.append(Spacer(1, 8))

    for k, v in avg_units.items():
        story.append(Paragraph(f"{k}: {v:.1f}%", normal))

        bar_width = int((v / 100) * 400)
        bar_color = get_bar_color(v)

        bar = Table([["", ""]],
                    colWidths=[bar_width, 400 - bar_width],
                    rowHeights=10)

        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))

        story.append(bar)
        story.append(Spacer(1, 6))

    # -----------------------------------------------------------
    # SAVE DOCUMENT
    # -----------------------------------------------------------
    save_path = macos_save_dialog(
        default_name="STAJ_Full_Report.pdf",
        prompt="Save Full PDF Report As"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"Full report saved at: {save_path}")
    else:
        print("Save cancelled.")

    # Remove temp gauge
    try:
        os.remove(gauge_path)
    except:
        pass




def commandbutton7_colored_report_old(ws):
    """
    Combined report: overall, outcome, and lead unit progress
    with gauge and horizontal colored bars
    """
    avgOverall, outcomes, avg_units = compute_all_progress(ws)

    if avgOverall is None:
        print("No data found to generate full report.")
        return

    # Draw gauge
    gauge_path = "gauge_dial.png"
    draw_color_coded_gauge(avgOverall, gauge_path)

    styles = getSampleStyleSheet()
    story = []

    # --- Title ---
    story.append(Paragraph("<b>STAJ IMPLEMENTATION FULL REPORT</b>", styles["Title"]))
    story.append(Spacer(1, 12))
    
  # Add summary lines
    story.append(Paragraph("This report provides a consolidated summary of the current status of STAJ implementation, based on progress records submitted by implementing units.\n", styles["Normal"]))
    story.append(Paragraph("It highlights key milestones achieved, overall performance, and priority areas requiring targeted action to strengthen implementation across the Judiciary.", styles["Normal"]))
    story.append(Spacer(1, 12))


    # --- Overall Achievement with Gauge side by side ---
    overall_data = [
        [
            Paragraph(f"<b>Overall Achievement: {avgOverall:.1f}%</b>", styles["Normal"]),
            Image(gauge_path, width=200, height=120)
        ]
    ]
    table = Table(overall_data, colWidths=[250, 250])
    table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER')
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph("="*75, styles["Normal"]))  # double line separator
    story.append(Spacer(1, 12))

    # --- Outcome Progress ---
    story.append(Paragraph("<b>Outcome Progress</b>", styles["Heading2"]))
    for k, v in outcomes.items():
        story.append(Paragraph(f"{k}: {v:.2f}%", styles["Normal"]))
        # Horizontal bar
        bar_width = int((v/100)*400)
        bar_color = get_bar_color(v)
        bar = Table([[ "", "" ]], colWidths=[bar_width, 400-bar_width], rowHeights=12)
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))
        story.append(bar)
        story.append(Spacer(1, 6))
    story.append(Spacer(1, 12))
    story.append(Paragraph("="*75, styles["Normal"]))
    story.append(Spacer(1, 12))

    # --- Lead Unit Progress ---
    story.append(Paragraph("<b>Lead Unit Progress</b>", styles["Heading2"]))
    for k, v in avg_units.items():
        story.append(Paragraph(f"{k}: {v:.1f}%", styles["Normal"]))
        bar_width = int((v/100)*400)
        bar_color = get_bar_color(v)
        bar = Table([[ "", "" ]], colWidths=[bar_width, 400-bar_width], rowHeights=12)
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))
        story.append(bar)
        story.append(Spacer(1, 6))

    # --- Build PDF ---
    save_path = macos_save_dialog(default_name="STAJ_Full_Report.pdf", prompt="Save Full PDF Report As")
    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story)
        print(f"Full report saved as PDF: {save_path}")
    else:
        print("Save cancelled.")

    # Cleanup
    try:
        os.remove(gauge_path)
    except Exception:
        pass



def commandbutton7_overall_and_outcome_and_unit_report(ws):
    last_row = 0
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=5).value not in (None, ""):
            last_row = r
            break
    if last_row == 0:
        print("No lead unit rows found.")
        return

    achievements = []
    dict_units_sum = {}
    dict_units_count = {}

    for i in range(2, last_row + 1):
        unit_cell = ws.cell(row=i, column=5).value
        if unit_cell in (None, ""):
            continue
        units = extract_units(str(unit_cell))
        target_value = ws.cell(row=i, column=8).value
        if not is_number(target_value) or float(target_value) == 0:
            continue
        target_value = float(target_value)
        progress = sum(
            float(ws.cell(row=i, column=c).value)
            for c in range(9, 21)
            if is_number(ws.cell(row=i, column=c).value)
        )
        row_pct = (progress / target_value) * 100

        # overall
        achievements.append(row_pct)

        # per lead unit
        for u in units:
            dict_units_sum[u] = dict_units_sum.get(u, 0.0) + row_pct
            dict_units_count[u] = dict_units_count.get(u, 0) + 1

    # Compute averages per unit
    avg_units = {k: dict_units_sum[k] / dict_units_count[k] for k in dict_units_sum}

    # overall achievement
    avgOverall = sum(achievements) / len(achievements) if achievements else 0.0

    # outcomes
    outcomes = compute_outcome_progress(ws)

    # build lines
    lines = []
    lines.append("Overall Progress Report:")
    lines.append(f"Overall Achievement: {avgOverall:.1f}%\n")
    lines.append("Outcome Progress Report:")
    for k, v in outcomes.items():
        lines.append(f"{k}: {v:.2f}%")
    lines.append("\nLead Unit Progress Report:")
    for k, v in avg_units.items():
        lines.append(f"{k}: {v:.1f}%")

    print("\n".join(lines))

def report_generator_all(ws):
    last_row_lead_units = 0
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=5).value not in (None, ""):
            last_row_lead_units = r
            break
    if last_row_lead_units == 0:
        print("No lead unit rows found.")
        return
    achievements = []
    dict_sum = {}
    dict_cnt = {}
    for i in range(2, last_row_lead_units + 1):
        unit_cell = ws.cell(row=i, column=5).value
        if unit_cell in (None, ""):
            continue
        units = extract_units(str(unit_cell))
        target_value = ws.cell(row=i, column=8).value
        if not is_number(target_value) or float(target_value) == 0:
            continue
        target_value = float(target_value)
        totalQuarterProgress = 0.0
        for c in range(9, 21):
            v = ws.cell(row=i, column=c).value
            if is_number(v):
                totalQuarterProgress += float(v)
        progress = (totalQuarterProgress / target_value) * 100 if target_value != 0 else 0.0
        achievements.append(progress)
        for u in units:
            dict_sum[u] = dict_sum.get(u, 0.0) + progress
            dict_cnt[u] = dict_cnt.get(u, 0) + 1
    avgOverall = sum(achievements) / len(achievements) if achievements else 0.0
    avg_units = {}
    for k, v in dict_sum.items():
        cnt = dict_cnt.get(k, 1)
        avg_units[k] = v / cnt if cnt != 0 else 0.0
    outcomes = compute_outcome_progress(ws)
    return avgOverall, outcomes, avg_units


from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle


def generate_full_pdf_report(ws):
    """
    Generates a structured PDF with:
    - Overall Achievement + Gauge
    - Outcome Progress with horizontal bars
    - Lead Unit Progress
    """
    avgOverall, outcomes, avg_units = compute_all_progress(ws)

    if avgOverall is None:
        print("No data found to generate full report.")
        return

    # Draw gauge
    gauge_path = "gauge_dial.png"
    draw_color_coded_gauge(avgOverall, gauge_path)

    styles = getSampleStyleSheet()
    story = []

    # --- Title ---
    story.append(Paragraph("<b>STAJ IMPLEMENTATION FULL REPORT</b>", styles["Title"]))
    story.append(Spacer(1, 12))

    # Add summary lines
    story.append(Paragraph("This report provides a consolidated summary of the current status of STAJ implementation, based on progress records submitted by implementing units.\n", styles["Normal"]))
    story.append(Paragraph("It highlights key milestones achieved, overall performance, and priority areas requiring targeted action to strengthen implementation across the Judiciary.", styles["Normal"]))
    story.append(Spacer(1, 12))
    
    # --- Overall Achievement with Gauge side by side ---
    overall_data = [
        [
            Paragraph(f"<b>Overall Achievement: {avgOverall:.1f}%</b>", styles["Normal"]),
            Image(gauge_path, width=200, height=120)
        ]
    ]
    table = Table(overall_data, colWidths=[250, 250])
    table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER')
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph("="*120, styles["Normal"]))  # double line separator
    story.append(Spacer(1, 12))

    # --- Outcome Progress ---
    story.append(Paragraph("<b>Outcome Progress</b>", styles["Heading2"]))
    for k, v in outcomes.items():
        story.append(Paragraph(f"{k}: {v:.2f}%", styles["Normal"]))
        # Create a horizontal bar
        bar_width = int((v/100)*400)  # 400px full width
        bar = Table([[ "", "" ]], colWidths=[bar_width, 400-bar_width], rowHeights=12)
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), colors.green),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))
        story.append(bar)
        story.append(Spacer(1, 6))
    story.append(Spacer(1, 12))
    story.append(Paragraph("="*120, styles["Normal"]))
    story.append(Spacer(1, 12))

    # --- Lead Unit Progress ---
    story.append(Paragraph("<b>Lead Unit Progress</b>", styles["Heading2"]))
    for k, v in avg_units.items():
        story.append(Paragraph(f"{k}: {v:.1f}%", styles["Normal"]))
        bar_width = int((v/100)*400)
        bar = Table([[ "", "" ]], colWidths=[bar_width, 400-bar_width], rowHeights=12)
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), colors.blue),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))
        story.append(bar)
        story.append(Spacer(1, 6))

    # --- Build PDF ---
    save_path = macos_save_dialog(default_name="STAJ_Full_Report.pdf", prompt="Save Full PDF Report As")
    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story)
        print(f"Full report saved as PDF: {save_path}")
    else:
        print("Save cancelled.")

    # Cleanup
    try:
        os.remove(gauge_path)
    except Exception:
        pass

from reportlab.lib import colors
from reportlab.lib import colors
'''
def get_bar_color(val):
    """Match gauge colors for progress bars"""
    if val <= 50:
        return colors.red
    elif val <= 75:
        return colors.orange  # amber
    elif val <= 94:
        return colors.lightgreen
    elif val <= 100:
        return colors.green
    else:
        return colors.darkgreen
'''
from reportlab.lib import colors
import math

def get_bar_color(val):
    """
    Returns a color matching the gauge's smooth transition:
    darkred → red → orangered → gold → lightgreen → green
    Ensures 0% is RED (never grey).
    """

    # Handle None or NaN explicitly
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return colors.HexColor("#EC3838")  # treat as 0% → red

    v = max(0, min(float(val), 100))  # clamp 0–100

    if v == 0:
        return colors.HexColor("#EC3838")  # explicit 0% safeguard
    elif v <= 15:
        return colors.HexColor("#EC3838")  # dark red
    elif v <= 30:
        return colors.HexColor("#C53333")  # red
    elif v <= 50:
        return colors.HexColor("#A23811")  # orangered
    elif v <= 75:
        return colors.HexColor("#FFD700")  # gold
    elif v <= 94:
        return colors.HexColor("#90EE90")  # light green
    else:
        return colors.HexColor("#008000")  # green


def generate_full_pdf_report_colored_bars(ws):
    """
    Generates a structured PDF with colored horizontal bars matching gauge logic
    """
    avgOverall, outcomes, avg_units = compute_all_progress(ws)

    if avgOverall is None:
        print("No data found to generate full report.")
        return

    # Draw gauge
    gauge_path = "gauge_dial.png"
    draw_color_coded_gauge(avgOverall, gauge_path)

    styles = getSampleStyleSheet()
    story = []

    # --- Title ---
    story.append(Paragraph("<b>STAJ IMPLEMENTATION FULL REPORT</b>", styles["Title"]))
    story.append(Spacer(1, 12))
    
    # Add summary lines
    story.append(Paragraph("This report provides a consolidated summary of the current status of STAJ implementation, based on progress records submitted by implementing units.\n", styles["Normal"]))
    story.append(Paragraph("It highlights key milestones achieved, overall performance, and priority areas requiring targeted action to strengthen implementation across the Judiciary.", styles["Normal"]))
    story.append(Spacer(1, 12))
    # --- Overall Achievement with Gauge side by side ---
    overall_data = [
        [
            Paragraph(f"<b>Overall Achievement: {avgOverall:.1f}%</b>", styles["Normal"]),
            Image(gauge_path, width=200, height=120)
        ]
    ]
    table = Table(overall_data, colWidths=[250, 250])
    table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER')
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph("="*120, styles["Normal"]))  # double line separator
    story.append(Spacer(1, 12))

    # --- Outcome Progress ---
    story.append(Paragraph("<b>Outcome Progress</b>", styles["Heading2"]))
    for k, v in outcomes.items():
        story.append(Paragraph(f"{k}: {v:.2f}%", styles["Normal"]))
        # Create a horizontal bar
        bar_width = int((v/100)*400)  # 400px full width
        bar_color = get_bar_color(v)
        bar = Table([[ "", "" ]], colWidths=[bar_width, 400-bar_width], rowHeights=12)
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))
        story.append(bar)
        story.append(Spacer(1, 6))
    story.append(Spacer(1, 12))
    story.append(Paragraph("="*120, styles["Normal"]))
    story.append(Spacer(1, 12))

    # --- Lead Unit Progress ---
    story.append(Paragraph("<b>Lead Unit Progress</b>", styles["Heading2"]))
    for k, v in avg_units.items():
        story.append(Paragraph(f"{k}: {v:.1f}%", styles["Normal"]))
        bar_width = int((v/100)*400)
        bar_color = get_bar_color(v)
        bar = Table([[ "", "" ]], colWidths=[bar_width, 400-bar_width], rowHeights=12)
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,0), bar_color),
            ('BACKGROUND', (1,0), (1,0), colors.lightgrey),
        ]))
        story.append(bar)
        story.append(Spacer(1, 6))

    # --- Build PDF ---
    save_path = macos_save_dialog(default_name="STAJ_Full_Report.pdf", prompt="Save Full PDF Report As")
    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story)
        print(f"Full report saved as PDF: {save_path}")
    else:
        print("Save cancelled.")

    # Cleanup
    try:
        os.remove(gauge_path)
    except Exception:
        pass
    

def generate_unit_and_output_progress(ws, target_col=8, qcols=list(range(9, 21))):
    """
    Generates implementing unit progress and output progress:
    Returns dict structure:
    {
        "UnitName": {
            "overall_progress": xx,
            "outputs": {
                "Output1": xx,
                "Output2": xx
            }
        }
    }
    """

    unit_progress = {}   # final dictionary

    for row in range(4, ws.max_row + 1):
        lead_unit = ws.cell(row=row, column=7).value
        output_name = ws.cell(row=row, column=3).value
        target = ws.cell(row=row, column=target_col).value

        if not lead_unit or not output_name or not target:
            continue  # skip empty or non-output rows

        # Collect quarter values
        q_values = []
        for col in qcols:
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)) and val >= 0:
                q_values.append(val)

        if not q_values:
            continue

        # Output progress = last quarter filled / target * 100
        last_val = q_values[-1]
        out_pct = (last_val / target) * 100 if target else 0

        if lead_unit not in unit_progress:
            unit_progress[lead_unit] = {
                "overall_progress": [],
                "outputs": {}
            }

        # store output progress
        unit_progress[lead_unit]["outputs"][output_name] = out_pct
        unit_progress[lead_unit]["overall_progress"].append(out_pct)

    # Final computation for each implementing unit
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        unit_progress[unit]["overall_progress"] = sum(arr) / len(arr) if arr else 0

    return unit_progress


# Each unit + output report



def commandbutton8_unit_output_report(ws):
    """
    Generates and prints Implementing Unit + Output progress summary
    """
    data = generate_unit_and_output_progress(ws)

    print("\nIMPLEMENTING UNIT + OUTPUT PROGRESS SUMMARY\n")
    for unit, info in data.items():
        print(f"\n--- {unit} ---")
        print(f"Overall Progress: {info['overall_progress']:.2f}%")

        for output_name, pct in info["outputs"].items():
            print(f"   • {output_name}: {pct:.2f}%")
            
        
def generate_unit_and_output_progress_groupedx(ws):
    """
    Generates implementing unit progress and output progress grouped by Lead Unit.
    Column mapping:
        Lead Unit: F (6)
        Output: C (3)
        Target: H (8)
        Quarters: I → T (9-20)
    Returns:
    {
        "LeadUnitName": {
            "overall_progress": xx,
            "outputs": {
                "Output1": xx,
                "Output2": xx
            }
        }
    }
    """
    unit_progress = {}

    for row in range(4, ws.max_row + 1):
        lead_unit = ws.cell(row=row, column=6).value  # F
        output_name = ws.cell(row=row, column=3).value  # C
        target = ws.cell(row=row, column=8).value  # H

        if not lead_unit or not output_name:
            continue

        # Ensure target is numeric
        try:
            target_val = float(target)
        except:
            target_val = None

        # Collect quarter values from I → T (columns 9 → 20)
        q_values = []
        for col in range(9, 21):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                q_values.append(val)
            elif isinstance(val, str) and val.strip().endswith('%'):
                try:
                    q_values.append(float(val.strip('%')))
                except:
                    pass
      
        # Sum all quarter values
        total_val = sum(q_values) if q_values else 0

# Compute output progress as cumulative percentage
        out_pct = (total_val / target_val * 100) if target_val else 0


        # CAP at 100%
        if out_pct > 100:
            out_pct = 100

        if lead_unit not in unit_progress:
            unit_progress[lead_unit] = {"overall_progress": [], "outputs": {}}

        unit_progress[lead_unit]["outputs"][output_name] = out_pct
        unit_progress[lead_unit]["overall_progress"].append(out_pct)

    # Compute overall progress per unit
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        data["overall_progress"] = sum(arr) / len(arr) if arr else 0

    return unit_progress

def generate_unit_and_output_progress_grouped_old(ws):
    """
    Generates implementing unit progress and output progress grouped by Lead Unit.
    Column mapping:
        Lead Unit: F (6)
        Output: C (3)
        Target: H (8)
        Quarters: I → T (9-20)
    Returns:
    {
        "LeadUnitName": {
            "overall_progress": xx,
            "outputs": {
                "Output1": xx,
                "Output2": xx
            }
        }
    }
    """
    unit_progress = {}

    for row in range(4, ws.max_row + 1):
        lead_unit = ws.cell(row=row, column=6).value  # F
        output_name = ws.cell(row=row, column=3).value  # C
        target = ws.cell(row=row, column=8).value  # H

        if not lead_unit or not output_name:
            continue

        # Ensure target is numeric
        try:
            target_val = float(target)
        except:
            target_val = None

        # Collect quarter values from I → T (columns 9 → 20)
        q_values = []
        for col in range(9, 21):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                q_values.append(val)
            elif isinstance(val, str) and val.strip().endswith('%'):
                try:
                    q_values.append(float(val.strip('%')))
                except:
                    pass

        # Cumulative progress across all quarters
        total_val = sum(q_values) if q_values else 0
        out_pct = (total_val / target_val * 100) if target_val else 0

        # CAP at 100%
        if out_pct > 100:
            out_pct = 100

        if lead_unit not in unit_progress:
            unit_progress[lead_unit] = {"overall_progress": [], "outputs": {}}

        unit_progress[lead_unit]["outputs"][output_name] = out_pct
        unit_progress[lead_unit]["overall_progress"].append(out_pct)

    # Compute overall progress per unit, capped at 100%
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        overall = sum(arr) / len(arr) if arr else 0
        if overall > 100:
            overall = 100
        data["overall_progress"] = overall

    return unit_progress

def generate_unit_and_output_progress_grouped_old2(ws):
    """
    Generates implementing unit progress and output progress grouped by Lead Unit.
    Column mapping:
        Lead Unit: F (6)
        Output: C (3)
        Target: H (8)
        Quarters: I → T (9-20)
    Returns:
    {
        "LeadUnitName": {
            "overall_progress": xx,
            "outputs": {
                "Output1": xx,
                "Output2": xx
            }
        }
    }
    """
    unit_progress = {}

    for row in range(4, ws.max_row + 1):
        lead_unit = ws.cell(row=row, column=6).value  # F
        output_name = ws.cell(row=row, column=3).value  # C
        target = ws.cell(row=row, column=8).value  # H

        if not lead_unit or not output_name or target in (None, "", "-"):
            continue  # skip if target is blank or non-numeric

        # Ensure target is numeric
        try:
            target_val = float(target)
        except:
            continue  # skip if target cannot be converted

        # Collect quarter values from I → T (columns 9 → 20)
        q_values = []
        for col in range(9, 21):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                q_values.append(val)
            elif isinstance(val, str) and val.strip().endswith('%'):
                try:
                    q_values.append(float(val.strip('%')))
                except:
                    pass

        # Skip if no quarter values
        if not q_values:
            continue

        # Cumulative progress across all quarters
        total_val = sum(q_values)
        out_pct = (total_val / target_val * 100) if target_val else 0

        # CAP at 100%
        if out_pct > 100:
            out_pct = 100

        if lead_unit not in unit_progress:
            unit_progress[lead_unit] = {"overall_progress": [], "outputs": {}}

        unit_progress[lead_unit]["outputs"][output_name] = out_pct
        unit_progress[lead_unit]["overall_progress"].append(out_pct)

    # Compute overall progress per unit, capped at 100%
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        overall = sum(arr) / len(arr) if arr else 0
        if overall > 100:
            overall = 100
        data["overall_progress"] = overall

    return unit_progress


from typing import List

from typing import List

REGISTRAR_UNITS = [
    "ORSC",     # Office of the Registrar Supreme Court
    "ORCOA",    # Office of the Registrar Court of Appeal
    "ORMC",     # Office of the Registrar Magistrates Courts
    "ORELRC",   # Office of the Registrar Employment & Labour Relations Court
    "ORELC",    # Office of the Registrar Environment & Land Court
    "ORT"       # Office of the Registrar Tribunals
]

def extract_units(unit_text: str) -> List[str]:
    """
    Splits a string containing multiple units into a list of clean unit names.
    Handles separators: ; / & and ' and '.
    If 'Registrar' or 'Registrars' is present, it is replaced with
    the standardized registrar units.
    """
    if not unit_text:
        return []

    s = unit_text.strip()

    # Normalize separators
    for sep in [';', '/', '&', ' and ']:
        s = s.replace(sep, ',')

    parts = [p.strip() for p in s.split(',') if p.strip()]

    output_units = []

    for part in parts:
        if part.lower() in {"registrar", "registrars"}:
            output_units.extend(REGISTRAR_UNITS)
        else:
            output_units.append(part)

    # Remove duplicates while preserving order
    seen = set()
    final_units = []
    for u in output_units:
        if u not in seen:
            seen.add(u)
            final_units.append(u)

    return final_units

#def extract_units(unit_text: str) -> List[str]:
    """
    Splits a string containing multiple units into a list of clean unit names.
    Handles separators: ; / & and ' and '
    """
 #   if not unit_text:
  #      return []
  #  s = unit_text.strip()
  #  for sep in [';', '/', '&', ' and ']:
  #      s = s.replace(sep, ',')
  #  parts = [p.strip() for p in s.split(',') if p.strip()]
  #  return parts


def generate_unit_and_output_progress_grouped(ws):
    """
    Generates implementing unit progress and output progress grouped by Lead Unit.
    Handles multiple units per row.
    """
    unit_progress = {}

    for row in range(4, ws.max_row + 1):

        # --------------------------------------------
        # Extract row values
        # --------------------------------------------
        lead_unit_text = ws.cell(row=row, column=6).value   # F
        output_name = ws.cell(row=row, column=3).value       # C
        uom = ws.cell(row=row, column=7).value               # G
        target = ws.cell(row=row, column=8).value            # H

        # Skip incomplete rows
        if not lead_unit_text or not output_name:
            continue
        if target in (None, "", "-", " "):
            continue

        # --------------------------------------------
        # Parse Unit of Measure
        # --------------------------------------------
        uom_text = str(uom).lower() if uom else ""

        is_percentage = ("percent" in uom_text) or ("%" in uom_text)

        # --------------------------------------------
        # Convert target correctly based on UoM
        # --------------------------------------------
        target_val = None

        # Target appears as number (e.g., 3) or string (e.g., "3", "50%")
        try:
            if isinstance(target, str) and target.strip().endswith("%"):
                # remove % and convert
                target_val = float(target.strip().replace("%", ""))
                is_percentage = True  # enforce
            else:
                target_val = float(target)
        except:
            continue  # skip non-numeric target

        # If UoM is percentage and value is decimal (e.g., 0.4), convert to % (40)
        if is_percentage and target_val <= 1:
            target_val *= 100

        # --------------------------------------------
        # Collect quarterly cumulative achievement
        # --------------------------------------------
        q_values = []

        for col in range(9, 21):  # I → T
            val = ws.cell(row=row, column=col).value

            if val is None or val == "":
                continue

            # ---- If UoM is percentage ----
            if is_percentage:
                # Case 1: 40% written as text
                if isinstance(val, str) and val.strip().endswith("%"):
                    try:
                        q_values.append(float(val.strip().replace("%", "")))
                    except:
                        pass

                # Case 2: decimal 0.40 meaning 40%
                elif isinstance(val, (int, float)):
                    if val <= 1:     # interpret as 0.4 → 40%
                        q_values.append(val * 100)
                    else:
                        q_values.append(val)

                # ignore non-numeric
                continue

            # ---- If UoM is a number ----
            else:
                # case number
                if isinstance(val, (int, float)):
                    q_values.append(float(val))
                    continue

                # case text number
                if isinstance(val, str):
                    try:
                        q_values.append(float(val))
                    except:
                        pass

        if not q_values:
            continue

        total_ach = sum(q_values)

        # --------------------------------------------
        # Compute progress percentage
        # --------------------------------------------
        if target_val != 0:
            out_pct = (total_ach / target_val) * 100
        else:
            out_pct = 0

        if out_pct > 100:
            out_pct = 100

        # --------------------------------------------
        # Split multiple lead units (e.g., ORSC & DCRJ)
        # --------------------------------------------
        lead_units = extract_units(lead_unit_text)

        for lead_unit in lead_units:
            if lead_unit not in unit_progress:
                unit_progress[lead_unit] = {
                    "overall_progress": [],
                    "outputs": {}
                }

            unit_progress[lead_unit]["outputs"][output_name] = {
                "target": target_val,
                "achievement": total_ach,
                "progress": out_pct,
                "uom": uom
            }

            unit_progress[lead_unit]["overall_progress"].append(out_pct)

    # --------------------------------------------
    # Finalize overall progress per implementing unit
    # --------------------------------------------
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        overall = sum(arr) / len(arr) if arr else 0
        data["overall_progress"] = min(overall, 100)

    return unit_progress

def generate_unit_and_output_progress_grouped_vcx(ws):
    """
    Generates implementing unit progress and output progress grouped by Lead Unit.
    Handles multiple units per row.
    
    Column mapping:
        Output: C (3)
        Unit of Measure: G (7)
        Lead Unit: F (6)
        Target: H (8)
        Quarters: I → T (9-20)

    Returns:
    {
        "LeadUnitName": {
            "overall_progress": xx,
            "outputs": {
                "Output1": {
                    "progress": xx,
                    "target": xx,
                    "achievement": xx
                }
            }
        }
    }
    """
    unit_progress = {}

    for row in range(4, ws.max_row + 1):

        # --------------------------------------------
        # Extract row values
        # --------------------------------------------
        lead_unit_text = ws.cell(row=row, column=6).value   # F
        output_name = ws.cell(row=row, column=3).value       # C
        uom = ws.cell(row=row, column=7).value               # G
        target = ws.cell(row=row, column=8).value            # H

        # Skip incomplete rows
        if not lead_unit_text or not output_name:
            continue
        if target in (None, "", "-", " "):
            continue

        # --------------------------------------------
        # Convert target based on Unit of Measure
        # --------------------------------------------
        try:
            target_val = float(target)
        except:
            continue

        # UoM fix: If unit is percentage and value is like "0.50" => treat as 50%
        if uom and isinstance(target_val, float):
            uom_text = str(uom).lower()

            if "percent" in uom_text or "%" in uom_text:
                #if target_val <= 1:        # then treat as decimal percentage
                    target_val *= 100

        # --------------------------------------------
        # Collect quarterly cumulative achievement
        # --------------------------------------------
        q_values = []
        for col in range(9, 21):  # I → T
            val = ws.cell(row=row, column=col).value

            if isinstance(val, (int, float)):
                q_values.append(val)
            elif isinstance(val, str) and val.strip().endswith('%'):
                try:
                    q_values.append(float(val.strip('%')))
                except:
                    pass

        if not q_values:
            continue

        total_ach = sum(q_values)

        # --------------------------------------------
        # Compute progress percentage
        # --------------------------------------------
        if target_val:
            out_pct = (total_ach / target_val) * 100
        else:
            out_pct = 0

        # Cap progress at 100%
        if out_pct > 100:
            out_pct = 100

        # --------------------------------------------
        # Split multiple lead units (ORSC & DCRJ → 2 units)
        # --------------------------------------------
        lead_units = extract_units(lead_unit_text)

        for lead_unit in lead_units:
            if lead_unit not in unit_progress:
                unit_progress[lead_unit] = {
                    "overall_progress": [],
                    "outputs": {}
                }

            # Store richer output data
            unit_progress[lead_unit]["outputs"][output_name] = {
                "target": target_val,
                "achievement": total_ach,
                "progress": out_pct,
                "uom": uom
            }

            # Add to unit-level overall average
            unit_progress[lead_unit]["overall_progress"].append(out_pct)

    # --------------------------------------------
    # Finalize overall progress per implementing unit
    # --------------------------------------------
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        overall = sum(arr) / len(arr) if arr else 0
        data["overall_progress"] = min(overall, 100)  # cap at 100%

    return unit_progress



def generate_unit_and_output_progress_groupedxxx(ws):
    """
    Generates implementing unit progress and output progress grouped by Lead Unit.
    Handles multiple units per row.
    Column mapping:
        Lead Unit: F (6)
        Output: C (3)
        Target: H (8)
        Quarters: I → T (9-20)
    Returns:
    {
        "LeadUnitName": {
            "overall_progress": xx,
            "outputs": {
                "Output1": xx,
                "Output2": xx
            }
        }
    }
    """
    unit_progress = {}

    for row in range(4, ws.max_row + 1):
        lead_unit_text = ws.cell(row=row, column=6).value  # F
        output_name = ws.cell(row=row, column=3).value  # C
        target = ws.cell(row=row, column=8).value  # H

        if not lead_unit_text or not output_name or target in (None, "", "-"):
            continue  # skip if target is blank or invalid

        try:
            target_val = float(target)
        except:
            continue

        # Collect quarter values from I → T (columns 9 → 20)
        q_values = []
        for col in range(9, 21):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                q_values.append(val)
            elif isinstance(val, str) and val.strip().endswith('%'):
                try:
                    q_values.append(float(val.strip('%')))
                except:
                    pass

        if not q_values:
            continue

        # Cumulative progress
        total_val = sum(q_values)
        out_pct = (total_val / target_val * 100) if target_val else 0
        if out_pct > 100:
            out_pct = 100

        # Extract multiple units
        lead_units = extract_units(lead_unit_text)

        for lead_unit in lead_units:
            if lead_unit not in unit_progress:
                unit_progress[lead_unit] = {"overall_progress": [], "outputs": {}}

            unit_progress[lead_unit]["outputs"][output_name] = out_pct
            unit_progress[lead_unit]["overall_progress"].append(out_pct)

    # Compute overall progress per unit
    for unit, data in unit_progress.items():
        arr = data["overall_progress"]
        overall = sum(arr) / len(arr) if arr else 0
        if overall > 100:
            overall = 100
        data["overall_progress"] = overall

    return unit_progress


def commandbutton8_unit_output_report_grouped(ws):
    """
    Print Implementing Unit + Output Progress grouped by Lead Unit
    """
    data = generate_unit_and_output_progress_grouped(ws)

    print("\nJUDICIARY OF KENYA")
    print("IMPLEMENTING UNIT PROGRESS REPORT\n")

    for unit in sorted(data.keys()):
        info = data[unit]
        print(f"Lead Unit: {unit} - Overall Progress: {info['overall_progress']:.2f}%")
        for output_name in sorted(info["outputs"].keys()):
            pct = info["outputs"][output_name]
            print(f"   • {output_name}: {pct:.2f}%")
        print("\n")


def generate_pdf_unit_output_report_old(ws):
    """
    Generates a PDF report for Implementing Unit + Output Progress
    grouped by Lead Unit, with progress bars.
    """
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import os

    data = generate_unit_and_output_progress_grouped(ws)

    if not data:
        print("No data available for Unit + Output report.")
        return

    # ------------------------
    # Styles
    # ------------------------
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=14,
        leading=20,
        alignment=1,
    )

    heading = ParagraphStyle(
        name="Heading2",
        parent=styles["Heading2"],
        fontName="Optima",
        fontSize=12,
        spaceAfter=2,
        bold=True,
        alignment=0,  # left
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=10,
        leading=14
    )

    # ------------------------
    # PDF Story
    # ------------------------
    story = []

    # Header
    story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", header_style))
    story.append(Spacer(1, 5))
    story.append(Paragraph("<b>IMPLEMENTING UNIT PROGRESS REPORT</b>", heading))
    story.append(Spacer(1, 10))

    # Iterate units
    for unit, info in data.items():
        story.append(Paragraph(f"<b>Lead Unit: {unit} - Overall Progress: {info['overall_progress']:.2f}%</b>", normal))
        story.append(Spacer(1, 3))

        for output_name, pct in info["outputs"].items():
            story.append(Paragraph(f"• {output_name}: {pct:.2f}%", normal))
            
            # Draw progress bar
            bar_width = int((pct / 100) * 400)
            bar_color = get_bar_color(pct)
            bar = Table(
                [["", ""]],
                colWidths=[bar_width, 400 - bar_width],
                rowHeights=10
            )
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), bar_color),
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
            ]))
            story.append(bar)
            story.append(Spacer(1, 5))
        
        story.append(Spacer(1, 10))

    # ------------------------
    # Save PDF
    # ------------------------
    save_path = macos_save_dialog(
        default_name="Unit_Output_Progress_Report.pdf",
        prompt="Save Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")


def generate_pdf_unit_output_report_grouped_o(ws):
    """
    Generates a PDF report for Implementing Unit + Output Progress
    grouped by Lead Unit, with progress bars.
    Column mapping:
        Lead Unit: F (6)
        Output: C (3)
        Target: H (8)
        Quarters: I → T (9-20)
    """
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import os

    data = generate_unit_and_output_progress_grouped_o(ws)
    if not data:
        print("No data available for Unit + Output report.")
        return

    # ------------------------
    # Styles
    # ------------------------
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=14,
        leading=20,
        alignment=1,  # Center
    )

    heading = ParagraphStyle(
        name="Heading2",
        parent=styles["Heading2"],
        fontName="Optima",
        fontSize=12,
        spaceAfter=2,
        alignment=0,  # Left
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=10,
        leading=14
    )

    # ------------------------
    # PDF Story
    # ------------------------
    story = []

    # Header
    story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", header_style))
    story.append(Spacer(1, 5))
    story.append(Paragraph("<b>IMPLEMENTING UNIT PROGRESS REPORT</b>", heading))
    story.append(Spacer(1, 10))

    # Iterate Lead Units
    for unit in sorted(data.keys()):
        info = data[unit]
        story.append(Paragraph(f"<b>Lead Unit: {unit} - Overall Progress: {info['overall_progress']:.2f}%</b>", normal))
        story.append(Spacer(1, 5))

        # Iterate outputs
        for output_name in sorted(info["outputs"].keys()):
            pct = info["outputs"][output_name]
            story.append(Paragraph(f"• {output_name}: {pct:.2f}%", normal))

            # Draw progress bar
            bar_width = int((pct / 100) * 400)
            bar_color = get_bar_color(pct)
            bar = Table(
                [["", ""]],
                colWidths=[bar_width, 400 - bar_width],
                rowHeights=10
            )
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), bar_color),
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
            ]))
            story.append(bar)
            story.append(Spacer(1, 5))

        story.append(Spacer(1, 10))

    # ------------------------
    # Save PDF
    # ------------------------
    save_path = macos_save_dialog(
        default_name="Unit_Output_Progress_Report.pdf",
        prompt="Save Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ------------------------
# PAGE HEADER + BORDER
# ------------------------
def draw_page_header_o(canvas, doc):
    """
    Draws the repeated header and page border on every page
    """
    canvas.saveState()

    # Draw thin border
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(0.5)
    margin = 15
    canvas.rect(
        margin,
        margin,
        doc.pagesize[0] - 2 * margin,
        doc.pagesize[1] - 2 * margin
    )

    # Header text
    canvas.setFont("Optima", 12)
    canvas.drawCentredString(doc.pagesize[0]/2, doc.pagesize[1] - 40, "JUDICIARY OF KENYA")
    canvas.drawCentredString(doc.pagesize[0]/2, doc.pagesize[1] - 60, "STAJ IMPLEMENTING UNIT PROGRESS REPORT")

    canvas.restoreState()


from reportlab.lib.units import mm

def draw_page_header(canvas, doc):
    canvas.saveState()

    width, height = doc.pagesize

    # ----------------------------
    # Draw Logo (centered)
    # ----------------------------
    logo_path = "/Users/jud-05/Desktop/nyambane/judiciary_logo.png"
    logo_width = 100
    logo_height = 60
    try:
        canvas.drawImage(
            logo_path,
            x=(width - logo_width)/2,
            y=height - 70,  # adjust from top
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True
        )
    except:
        # fallback text if logo not found
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(width/2, height - 50, "JUDICIARY")

    # ----------------------------
    # Draw Header Texts
    # ----------------------------
    canvas.setFont("Optima", 10)  # mini_header_style
    canvas.drawCentredString(width/2, height - 80, "JUDICIARY OF KENYA")

    canvas.setFont("Optima", 12)  # header_style
    canvas.drawCentredString(width/2, height - 100, "STAJ IMPLEMENTATION FULL REPORT")

    canvas.restoreState()


# ------------------------
# PDF function
# ------------------------
def generate_pdf_unit_output_report_grouped_old(ws):
    """
    Generates a PDF report for Implementing Unit + Output Progress
    grouped by Lead Unit, each on its own page,
    with repeated header and progress bars.
    """
    data = generate_unit_and_output_progress_grouped(ws)

    if not data:
        print("No data available for Unit + Output report.")
        return

    # Styles
    styles = getSampleStyleSheet()
    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=10,
        leading=14
    )

    story = []

    for unit, info in data.items():
        # Lead Unit title
        story.append(Spacer(1, 70))  # leave space for header
        story.append(Paragraph(f"<b>{unit} - Overall Progress: {info['overall_progress']:.0f}%</b>", normal))
        story.append(Spacer(1, 5))

        # Output progress
        for output_name, pct in info["outputs"].items():
            story.append(Paragraph(f"• {output_name}: {pct:.0f}%", normal))

            # Draw progress bar
            bar_width = int((pct / 100) * 400)
            bar_color = get_bar_color(pct)
            bar = Table(
                [["", ""]],
                colWidths=[bar_width, 400 - bar_width],
                rowHeights=10
            )
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), bar_color),
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
            ]))
            story.append(bar)
            story.append(Spacer(1, 5))

        # Page break after each unit
        story.append(PageBreak())

    # Save PDF
    save_path = macos_save_dialog(
        default_name="Unit_Output_Progress_Report_Per_Unit.pdf",
        prompt="Save Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_header, onLaterPages=draw_page_header)
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")


def generate_pdf_unit_output_report_with_target_old(ws):
    """
    Generates a PDF report for Implementing Unit + Output Progress
    grouped by Lead Unit, showing Target, Achievement, and Progress %.
    Each unit starts on a new page.
    """
    data = generate_unit_and_output_progress_grouped(ws)
    if not data:
        print("No data available for Unit + Output report.")
        return

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=12,
        leading=20,
        alignment=1,  # Center
   )

    heading = ParagraphStyle(
        name="Heading2",
        parent=styles["Heading2"],
        fontName="Optima",
        fontSize=10,
        spaceAfter=2,
        bold=True,
        alignment=0,  # Left
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=8,
        leading=14
    )

    story = []

    # Logo
    try:
        from reportlab.platypus import Image
        logo = Image("/Users/jud-05/Desktop/nyambane/judiciary_logo.png", width=100, height=60)
    except:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)
    logo.hAlign = "CENTER"

    for unit, info in data.items():
        # Start new page per unit
        story.append(PageBreak())
        story.append(logo)
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", header_style))
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>STAJ IMPLEMENTING UNIT PROGRESS REPORT</b>", header_style))
        story.append(Spacer(1, 10))

        # Unit title
        #story.append(Paragraph(f"<b>{unit} - Overall Progress: {info['overall_progress']:.2f}%</b>", header_style))
        story.append(
    Paragraph(
        f"<font color='blue'><b>{unit} - Overall Progress: {info['overall_progress']:.2f}%</b></font>",
        header_style
    )
)

        story.append(Spacer(1, 5))
        
        for idx, (output_name, out_data) in enumerate(info["outputs"].items(), start=1):
    # Expect out_data as dict: {'progress': xx, 'target': yy, 'achievement': zz}
            pct = out_data.get('progress', 0)
            target_val = int(out_data.get("target", 0))
            #target_val = out_data.get('target', '-')
            achievement = int(out_data.get('achievement', 0))
    
    # Paragraph text with numbering
            story.append(Paragraph(
            f"{idx}. {output_name}: Target {target_val}; Achievement {achievement}; Progress {pct:.0f}%",
            normal
    ))

            # Progress bar
            bar_width = int((pct / 100) * 400)
            bar_color = get_bar_color(pct)
            bar = Table(
                [["", ""]],
                colWidths=[bar_width, 400 - bar_width],
                rowHeights=10
            )
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), bar_color),
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
            ]))
            story.append(bar)
            story.append(Spacer(1, 5))

        story.append(Spacer(1, 10))

    save_path = macos_save_dialog(
        default_name="Unit_Output_Progress_Report_Target.pdf",
        prompt="Save Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")

from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4


def get_bar_color(val):
    """Returns bar color; ensures 0% is red, never grey."""
    v = max(0, min(float(val), 100))
    if v <= 15:
        return colors.HexColor("#EC3838")
    elif v <= 30:
        return colors.HexColor("#C53333")
    elif v <= 50:
        return colors.HexColor("#A23811")
    elif v <= 75:
        return colors.HexColor("#FFD700")
    elif v <= 94:
        return colors.HexColor("#90EE90")
    else:
        return colors.HexColor("#008000")


def generate_pdf_unit_output_report_with_target(ws):
    """
    Generates a PDF report for Implementing Unit + Output Progress
    grouped by Lead Unit, showing Target, Achievement, and Progress %.
    Each unit starts on a new page.
    Outputs are sorted from highest → lowest progress.
    """
    data = generate_unit_and_output_progress_grouped(ws)
    if not data:
        print("No data available for Unit + Output report.")
        return

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=12,
        leading=20,
        alignment=1  # Center
    )
    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=8,
        leading=14
    )

    story = []

    # Logo
    try:
        logo = Image("/Users/jud-05/Desktop/nyambane/judiciary_logo.png", width=100, height=60)
    except Exception:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)
    logo.hAlign = "CENTER"

    rendered_units = 0  # Track first page

    for unit, info in data.items():
        # Add page break only after first unit
        if rendered_units > 0:
            story.append(PageBreak())
        rendered_units += 1

        # Header
        story.append(logo)
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", header_style))
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>STAJ IMPLEMENTING UNIT PROGRESS REPORT</b>", header_style))
        story.append(Spacer(1, 10))
        story.append(
            Paragraph(
                f"<font color='blue'><b>{unit} - Overall Progress: {info['overall_progress']:.2f}%</b></font>",
                header_style
            )
        )
        story.append(Spacer(1, 5))

        # Sort outputs descending by progress
        sorted_outputs = sorted(
            info["outputs"].items(),
            key=lambda item: item[1].get("progress", 0),
            reverse=True
        )

        for idx, (output_name, out_data) in enumerate(sorted_outputs, start=1):
            pct = out_data.get('progress', 0)
            target_val = int(out_data.get("target", 0))
            achievement = int(out_data.get('achievement', 0))

            # Output paragraph
            story.append(Paragraph(
                f"{idx}. {output_name}: Target {target_val}; Achievement {achievement}; Progress {pct:.0f}%",
                normal
            ))

            # Progress bar
            total_width = 400
            pct = max(0, min(pct, 100))
            bar_width = int((pct / 100) * total_width)
            if pct == 0:
                bar_width = 3  # visible for 0%

            bar_color = get_bar_color(pct)
            bar = Table(
                [["", ""]],
                colWidths=[bar_width, total_width - bar_width],
                rowHeights=10
            )
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), bar_color),
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
            ]))
            story.append(bar)
            story.append(Spacer(1, 5))

        story.append(Spacer(1, 10))

    # Save PDF
    save_path = macos_save_dialog(
        default_name="Unit_Output_Progress_Report_Target.pdf",
        prompt="Save Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(story, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")
        
def generate_pdf_specific_units_output_report_with_target_old(ws):
    """
    Generates a PDF report for user-selected Implementing Unit(s) + Output Progress.
    User is prompted to enter unit name(s), e.g. OCJ or OCJ, DSPOP.
    Each selected unit starts on a new page (no blank pages).
    """

    # Prompt user to enter unit(s)
    units_input = input(
        "Enter Implementing Unit(s) separated by commas (e.g. OCJ or OCJ, DSPOP): "
    ).strip()

    if not units_input:
        print("No unit entered. Operation cancelled.")
        return

    selected_units = [u.strip().upper() for u in units_input.split(",")]

    data = generate_unit_and_output_progress_grouped(ws)
    if not data:
        print("No data available for Unit + Output report.")
        return

    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=12,
        leading=20,
        alignment=1,  # Center
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=8,
        leading=14
    )

    story = []

    # Logo
    try:
        from reportlab.platypus import Image
        logo = Image(
            "/Users/jud-05/Desktop/nyambane/judiciary_logo.png",
            width=100,
            height=60
        )
    except Exception:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)

    logo.hAlign = "CENTER"

    rendered_units = 0

    for unit, info in data.items():
        if unit.upper() not in selected_units:
            continue

        # Page break ONLY between rendered units
        if rendered_units > 0:
            story.append(PageBreak())

        rendered_units += 1

        story.append(logo)
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", header_style))
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>STAJ IMPLEMENTING UNIT PROGRESS REPORT</b>", header_style))
        story.append(Spacer(1, 10))

        story.append(
            Paragraph(
                f"<font color='blue'><b>{unit} - Overall Progress: "
                f"{info['overall_progress']:.2f}%</b></font>",
                header_style
            )
        )

        story.append(Spacer(1, 5))

        for idx, (output_name, out_data) in enumerate(info["outputs"].items(), start=1):
            pct = out_data.get("progress", 0)
            target_val = int(out_data.get("target", 0))

            #target_val = out_data.get("target", "-")
            achievement = int(out_data.get("achievement", 0))

            story.append(Paragraph(
                f"{idx}. {output_name}: Target {target_val}; "
                f"Achievement {achievement}; Progress {pct:.0f}%",
                normal
            ))
            total_width = 400
            bar_width = int((pct / 100) * total_width)
            # Force visibility for 0%
            if pct == 0:
                bar_width = 3  # small but visible red bar

            #bar_width = int((pct / 100) * 400)
            bar_color = get_bar_color(pct)

            bar = Table(
                [["", ""]],
                colWidths=[bar_width, 400 - bar_width],
                rowHeights=10
            )
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), bar_color),
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
            ]))

            story.append(bar)
            story.append(Spacer(1, 5))

        story.append(Spacer(1, 10))

    if rendered_units == 0:
        print(f"No matching units found for: {', '.join(selected_units)}")
        return

    # Dynamic file name based on selected units
    unit_suffix = "_".join(selected_units)
    file_name = f"{unit_suffix}_STAJ_Progress_Report.pdf"

    save_path = macos_save_dialog(
        default_name=file_name,
        prompt="Save Specific Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(
            story,
            onFirstPage=draw_page_border,
            onLaterPages=draw_page_border
        )
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")



def generate_pdf_specific_units_output_report_with_target(ws):
    """
    Generates a PDF report for user-selected Implementing Unit(s) + Output Progress.
    User is prompted to enter unit name(s), e.g. OCJ or OCJ, DSPOP.
    Each selected unit starts on a new page (no blank pages).
    """

    # Prompt user to enter unit(s)
    units_input = input(
        "Enter Implementing Unit(s) separated by commas (e.g. OCJ or OCJ, DSPOP): "
    ).strip()

    if not units_input:
        print("No unit entered. Operation cancelled.")
        return

    selected_units = [u.strip().upper() for u in units_input.split(",")]

    data = generate_unit_and_output_progress_grouped(ws)
    if not data:
        print("No data available for Unit + Output report.")
        return

    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        name="Header",
        fontName="Optima",
        fontSize=12,
        leading=20,
        alignment=1,  # Center
    )

    normal = ParagraphStyle(
        name="NormalOptima",
        parent=styles["Normal"],
        fontName="Optima",
        fontSize=8,
        leading=14
    )

    story = []

    # Logo
    try:
        from reportlab.platypus import Image
        logo = Image(
            "/Users/jud-05/Desktop/nyambane/judiciary_logo.png",
            width=100,
            height=60
        )
    except Exception:
        logo = Paragraph("<b>JUDICIARY</b>", header_style)

    logo.hAlign = "CENTER"

    rendered_units = 0

    for unit, info in data.items():
        if unit.upper() not in selected_units:
            continue

        # Page break ONLY between rendered units
        if rendered_units > 0:
            story.append(PageBreak())

        rendered_units += 1

        story.append(logo)
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>JUDICIARY OF KENYA</b>", header_style))
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>STAJ IMPLEMENTING UNIT PROGRESS REPORT</b>", header_style))
        story.append(Spacer(1, 10))

        story.append(
            Paragraph(
                f"<font color='blue'><b>{unit} - Overall Progress: "
                f"{info['overall_progress']:.2f}%</b></font>",
                header_style
            )
        )

        story.append(Spacer(1, 5))

        # 🔹 SORT outputs by progress (highest → lowest)
        sorted_outputs = sorted(
            info["outputs"].items(),
            key=lambda item: item[1].get("progress", 0),
            reverse=True
        )

        for idx, (output_name, out_data) in enumerate(sorted_outputs, start=1):
            pct = out_data.get("progress", 0)
            target_val = int(out_data.get("target", 0))
            achievement = int(out_data.get("achievement", 0))

            story.append(Paragraph(
                f"{idx}. {output_name}: Target {target_val}; "
                f"Achievement {achievement}; Progress {pct:.0f}%",
                normal
            ))

            total_width = 400
            bar_width = int((pct / 100) * total_width)

            # Force visibility for 0%
            if pct == 0:
                bar_width = 3

            bar_color = get_bar_color(pct)

            bar = Table(
                [["", ""]],
                colWidths=[bar_width, total_width - bar_width],
                rowHeights=10
            )
            bar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), bar_color),
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
            ]))

            story.append(bar)
            story.append(Spacer(1, 5))

        story.append(Spacer(1, 10))

    if rendered_units == 0:
        print(f"No matching units found for: {', '.join(selected_units)}")
        return

    # Dynamic file name based on selected units
    unit_suffix = "_".join(selected_units)
    file_name = f"{unit_suffix}_STAJ_Progress_Report.pdf"

    save_path = macos_save_dialog(
        default_name=file_name,
        prompt="Save Specific Unit + Output Progress PDF"
    )

    if save_path:
        doc = SimpleDocTemplate(save_path, pagesize=A4)
        doc.build(
            story,
            onFirstPage=draw_page_border,
            onLaterPages=draw_page_border
        )
        print(f"PDF report saved at: {save_path}")
    else:
        print("Save cancelled.")


   # -------------------------
# Main menu
# -------------------------

def main_menu():
    print("STAJ Tools - Python Conversion (macOS-safe)")
    print("Choose an action:")
    #print("1 → CommandButton1: Generate gauge PDF")
    #print("2 → CommandButton3: Outcome Progress Report")
    #print("4 → CommandButton6: Unit progress summary")
    print("5 → CommandButton7: Combined Overall/Outcome/Lead Unit Report")
    #print("6 → Full PDF Report (Overall + Outcomes + Lead Units)")
    #print("7 → Implementing Unit + Output Progress Report (NEW)")
    print("8 → Full PDF Report including Unit + Output Progress (NEW)")  # ← added
    print("9-generate_pdf_specific_units_output_report_with_target")

    choice = input("Enter number (or q to quit): ").strip().lower()
    if choice == "q":
        return

    wb_path = macos_choose_file(prompt="Select STAJ Monitoring Tool workbook (Excel)")
    if not wb_path:
        print("No workbook selected. Exiting.")
        return

    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
    except Exception as e:
        print(f"Failed to open workbook: {e}")
        return

    if "STAJ Monitoring Tool" not in wb.sheetnames:
        print("Workbook does not contain sheet named 'STAJ Monitoring Tool'. Exiting.")
        return

    ws = wb["STAJ Monitoring Tool"]

    #if choice == "1":
     #   commandbutton1_generate_gauge_pdf(ws)
    #if choice == "2":
     #   commandbutton3_outcome_report(ws)
    #if choice == "4":
     #   commandbutton6_unit_summary(ws)
    if choice == "5":
        commandbutton7_colored_reportYY(ws)
   #elif choice == "6":
    #    generate_full_pdf_report(ws)
    #elif choice == "7":
     #   commandbutton8_unit_output_report(ws)
    elif choice == "8":
        generate_pdf_unit_output_report_with_target(ws)  # ← NEW FUNCTION TO CREATE
    elif choice =="9":
        generate_pdf_specific_units_output_report_with_target(ws)
    else:
        print("Unknown choice.")
    print("\nDone.")

if __name__ == "__main__":
    main_menu()
